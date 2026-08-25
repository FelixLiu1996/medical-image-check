from __future__ import annotations

import math
from bisect import bisect_left
from collections import Counter, defaultdict, deque
from collections.abc import Callable
from dataclasses import dataclass
from decimal import Decimal
from itertools import combinations

from openpyxl.utils import get_column_letter

from medical_image_check.domain.excel_settings import ExcelAnalysisSettings, decimal_text
from medical_image_check.domain.models import (
    EvidenceLocation,
    Finding,
    FindingType,
    RiskLevel,
    deterministic_finding_id,
)
from medical_image_check.infrastructure.spreadsheets import NumericCell

CELL_OPERATION_RULE_ID = "excel.cell.target_operation"
SERIES_SHUFFLED_RULE_ID = "excel.series.shuffled"
SERIES_FRAGMENT_RULE_ID = "excel.series.fragment_exact"
SERIES_NEAR_DUPLICATE_RULE_ID = "excel.series.near_duplicate"
SERIES_LINEAR_RULE_ID = "excel.series.linear"
REGION_EXACT_RULE_ID = "excel.region.exact"
STATISTICS_RULE_ID = "excel.series.statistics"

MINIMUM_SERIES_LENGTH = 3
MINIMUM_NEAR_DUPLICATE_LENGTH = 5
MINIMUM_STATISTICS_LENGTH = 8
MAX_FINDINGS_PER_RULE = 300
MAX_GROUPS_PER_BUCKET = 64
ALL_PAIRS_SERIES_LIMIT = 250


@dataclass(frozen=True, slots=True)
class _Series:
    source_path: str
    sheet: str
    column: int
    cells: tuple[NumericCell, ...]

    @property
    def values(self) -> tuple[Decimal, ...]:
        return tuple(Decimal(cell.canonical_value) for cell in self.cells)

    @property
    def location(self) -> EvidenceLocation:
        column = get_column_letter(self.column)
        return EvidenceLocation(
            self.source_path,
            self.sheet,
            f"{column}{self.cells[0].row}:{column}{self.cells[-1].row}",
            self.cells[0].hidden_sheet,
        )


def find_excel_pattern_findings(
    cells: list[NumericCell],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None = None,
) -> list[Finding]:
    series = _collect_series(cells)
    phases = (
        lambda: _find_cell_operations(cells, settings, checkpoint),
        lambda: _find_exact_series_fragments(series, settings, checkpoint),
        lambda: _find_shuffled_series(series, settings, checkpoint),
        lambda: _find_near_duplicate_series(series, settings, checkpoint),
        lambda: _find_robust_linear_series(series, settings, checkpoint),
        lambda: _find_exact_regions(cells, settings, checkpoint),
        lambda: _find_statistical_similarity(series, settings, checkpoint),
    )
    findings: list[Finding] = []
    for phase in phases:
        if checkpoint:
            checkpoint()
        findings.extend(phase())
    return findings


def _collect_series(cells: list[NumericCell]) -> list[_Series]:
    grouped: dict[tuple[str, str, int], list[NumericCell]] = defaultdict(list)
    for cell in cells:
        grouped[(cell.source_path, cell.sheet, cell.column)].append(cell)
    return [
        _Series(key[0], key[1], key[2], tuple(sorted(items, key=lambda cell: cell.row)))
        for key, items in sorted(grouped.items())
        if len(items) >= MINIMUM_SERIES_LENGTH
    ]


def _find_cell_operations(
    cells: list[NumericCell],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    grouped: dict[str, list[NumericCell]] = defaultdict(list)
    for cell in cells:
        grouped[cell.canonical_value].append(cell)
    values = {canonical: Decimal(canonical) for canonical in grouped}
    ordered = sorted(values)
    ordered_values = sorted((value, key) for key, value in values.items())
    seen: set[tuple[str, str, str]] = set()
    findings: list[Finding] = []

    def add_candidate(
        first_key: str,
        second_key: str,
        operation: str,
        target: Decimal,
        result: Decimal,
        *,
        symmetric: bool,
    ) -> None:
        if len(findings) >= MAX_FINDINGS_PER_RULE or first_key == second_key:
            return
        first_value = values[first_key]
        second_value = values[second_key]
        if _is_trivial_operation(first_value, second_value, operation, target):
            return
        pair = tuple(sorted((first_key, second_key))) if symmetric else (first_key, second_key)
        dedupe_key = (operation, pair[0], pair[1])
        if dedupe_key in seen:
            return
        seen.add(dedupe_key)
        first_cell = grouped[first_key][0]
        second_cell = grouped[second_key][0]
        locations = (first_cell.location, second_cell.location)
        symbol = {"add": "+", "subtract": "-", "multiply": "×", "divide": "÷"}[operation]
        findings.append(
            Finding(
                deterministic_finding_id(f"{CELL_OPERATION_RULE_ID}.{operation}", locations),
                CELL_OPERATION_RULE_ID,
                FindingType.SUSPECTED_REUSE,
                RiskLevel.LOW,
                "两个完整数值运算得到目标值",
                f"{first_key} {symbol} {second_key} 约等于目标 {decimal_text(target)}。",
                locations,
                0.55,
                {
                    "operation": operation,
                    "parameter": decimal_text(target),
                    "relation_result": decimal_text(result),
                    "matched_count": 1,
                    "alignment": "不同单元格中的完整真实值",
                    "first_series": _single_cell_evidence(first_cell),
                    "second_series": _single_cell_evidence(second_cell),
                    "paired_values": [
                        {
                            "position": 1,
                            "first_coordinate": first_cell.coordinate,
                            "first_value": first_key,
                            "second_coordinate": second_cell.coordinate,
                            "second_value": second_key,
                            "relation_result": decimal_text(result),
                        }
                    ],
                },
            )
        )

    for target_index, target in enumerate(settings.operation_targets):
        if checkpoint and target_index % 4 == 0:
            checkpoint()
        for first_index, first_key in enumerate(ordered):
            if len(findings) >= MAX_FINDINGS_PER_RULE:
                return findings
            first_value = values[first_key]
            for operation, expected, symmetric in (
                ("add", target - first_value, True),
                ("subtract", first_value - target, False),
                (
                    "multiply",
                    target / first_value if first_value != 0 else None,
                    True,
                ),
                ("divide", first_value / target if target != 0 else None, False),
            ):
                if expected is None:
                    continue
                second_key = _nearest_value_key(expected, ordered_values)
                if second_key is None:
                    continue
                second_value = values[second_key]
                result = _apply_operation(first_value, second_value, operation)
                if result is not None and settings.close(result, target):
                    add_candidate(
                        first_key,
                        second_key,
                        operation,
                        target,
                        result,
                        symmetric=symmetric,
                    )
            if first_index % 512 == 0 and checkpoint:
                checkpoint()
    return findings


def _nearest_value_key(expected: Decimal, ordered_values: list[tuple[Decimal, str]]) -> str | None:
    if not ordered_values:
        return None
    position = bisect_left(ordered_values, (expected, ""))
    candidates = ordered_values[max(0, position - 1) : position + 1]
    return min(candidates, key=lambda item: (abs(item[0] - expected), item[1]))[1]


def _apply_operation(first: Decimal, second: Decimal, operation: str) -> Decimal | None:
    if operation == "add":
        return first + second
    if operation == "subtract":
        return first - second
    if operation == "multiply":
        return first * second
    if second != 0:
        return first / second
    return None


def _is_trivial_operation(first: Decimal, second: Decimal, operation: str, target: Decimal) -> bool:
    if first in {Decimal(0), Decimal(1)} and second in {Decimal(0), Decimal(1)}:
        return True
    if operation == "add" and (first == 0 or second == 0):
        return True
    if operation == "multiply" and (first in {0, 1} or second in {0, 1}):
        return True
    if operation == "divide" and second == 1:
        return True
    return operation == "subtract" and second == 0 and first == target


def _find_exact_series_fragments(
    series: list[_Series],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    minimum_length = settings.medium_run_length
    signatures: dict[tuple[str, ...], list[tuple[int, int]]] = defaultdict(list)
    for series_index, item in enumerate(series):
        values = tuple(cell.canonical_value for cell in item.cells)
        for start in range(len(values) - minimum_length + 1):
            signature = values[start : start + minimum_length]
            if len(signatures[signature]) < MAX_GROUPS_PER_BUCKET:
                signatures[signature].append((series_index, start))

    candidates: dict[tuple[int, int, int, int, int], int] = {}
    covered_diagonals: dict[tuple[int, int, int], list[tuple[int, int]]] = defaultdict(list)
    for bucket_index, occurrences in enumerate(signatures.values()):
        if checkpoint and bucket_index % 128 == 0:
            checkpoint()
        for (first_index, first_start), (second_index, second_start) in combinations(
            occurrences, 2
        ):
            if first_index == second_index and abs(first_start - second_start) < minimum_length:
                continue
            diagonal = (first_index, second_index, second_start - first_start)
            if any(
                covered_start <= first_start < covered_end
                for covered_start, covered_end in covered_diagonals[diagonal]
            ):
                continue
            first = series[first_index]
            second = series[second_index]
            left = 0
            while (
                first_start - left > 0
                and second_start - left > 0
                and first.cells[first_start - left - 1].canonical_value
                == second.cells[second_start - left - 1].canonical_value
            ):
                left += 1
            first_begin = first_start - left
            second_begin = second_start - left
            length = minimum_length + left
            while (
                first_begin + length < len(first.cells)
                and second_begin + length < len(second.cells)
                and first.cells[first_begin + length].canonical_value
                == second.cells[second_begin + length].canonical_value
            ):
                length += 1
            covered_diagonals[diagonal].append((first_begin, first_begin + length))
            if first_index == second_index and _ranges_overlap(first_begin, second_begin, length):
                continue
            if (
                first_begin == 0
                and second_begin == 0
                and length == len(first.cells)
                and length == len(second.cells)
            ):
                continue
            candidates[(first_index, first_begin, second_index, second_begin, length)] = length

    findings: list[Finding] = []
    for key, length in sorted(candidates.items(), key=lambda item: (-item[1], item[0])):
        first_index, first_begin, second_index, second_begin, _ = key
        first = series[first_index]
        second = series[second_index]
        first_cells = first.cells[first_begin : first_begin + length]
        second_cells = second.cells[second_begin : second_begin + length]
        locations = (
            _series_range_location(first, first_cells),
            _series_range_location(second, second_cells),
        )
        findings.append(
            Finding(
                deterministic_finding_id(SERIES_FRAGMENT_RULE_ID, locations),
                SERIES_FRAGMENT_RULE_ID,
                FindingType.EXACT_DUPLICATE,
                settings.risk_for_run(length),
                "连续数值片段完全重复",
                f"两段连续数值序列有 {length} 个完整值按相同顺序重复。",
                locations,
                0.96,
                {
                    "matched_count": length,
                    "alignment": "按每列数值单元格出现顺序的连续片段",
                    "first_series": _series_slice_evidence(first, first_cells),
                    "second_series": _series_slice_evidence(second, second_cells),
                    "paired_values": [
                        {
                            "position": position,
                            "first_coordinate": left_cell.coordinate,
                            "first_value": left_cell.canonical_value,
                            "second_coordinate": right_cell.coordinate,
                            "second_value": right_cell.canonical_value,
                            "relation_result": "完整值相同",
                        }
                        for position, (left_cell, right_cell) in enumerate(
                            zip(first_cells, second_cells, strict=True), 1
                        )
                    ],
                },
            )
        )
        if len(findings) >= MAX_FINDINGS_PER_RULE:
            break
    return findings


def _ranges_overlap(first_start: int, second_start: int, length: int) -> bool:
    return not (first_start + length <= second_start or second_start + length <= first_start)


def _find_shuffled_series(
    series: list[_Series],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    buckets: dict[tuple[int, tuple[str, ...]], list[_Series]] = defaultdict(list)
    for item in series:
        signature = (len(item.cells), tuple(sorted(cell.canonical_value for cell in item.cells)))
        if len(buckets[signature]) < MAX_GROUPS_PER_BUCKET:
            buckets[signature].append(item)
    findings: list[Finding] = []
    for bucket_index, bucket in enumerate(buckets.values()):
        if checkpoint and bucket_index % 32 == 0:
            checkpoint()
        for first, second in combinations(bucket, 2):
            first_values = [cell.canonical_value for cell in first.cells]
            second_values = [cell.canonical_value for cell in second.cells]
            if first_values == second_values:
                continue
            paired = _pair_by_value(first, second)
            locations = (first.location, second.location)
            risk = settings.risk_for_run(len(first.cells))
            findings.append(
                Finding(
                    deterministic_finding_id(SERIES_SHUFFLED_RULE_ID, locations),
                    SERIES_SHUFFLED_RULE_ID,
                    FindingType.SUSPECTED_REUSE,
                    risk,
                    "同一组完整数值以不同顺序出现",
                    "两列包含完全相同的完整数值及出现次数，但排列顺序不同。",
                    locations,
                    0.9,
                    {
                        "matched_count": len(first.cells),
                        "order_changed_count": sum(
                            left != right
                            for left, right in zip(first_values, second_values, strict=True)
                        ),
                        "alignment": "按完整值及重复次数重新配对",
                        "first_series": _series_evidence(first),
                        "second_series": _series_evidence(second),
                        "paired_values": paired,
                    },
                )
            )
            if len(findings) >= MAX_FINDINGS_PER_RULE:
                return findings
    return findings


def _pair_by_value(first: _Series, second: _Series) -> list[dict[str, object]]:
    second_cells: dict[str, deque[NumericCell]] = defaultdict(deque)
    for cell in second.cells:
        second_cells[cell.canonical_value].append(cell)
    return [
        {
            "position": index,
            "first_coordinate": cell.coordinate,
            "first_value": cell.canonical_value,
            "second_coordinate": matched.coordinate,
            "second_value": matched.canonical_value,
            "relation_result": "完整值相同",
        }
        for index, cell in enumerate(first.cells, start=1)
        for matched in (second_cells[cell.canonical_value].popleft(),)
    ]


def _find_near_duplicate_series(
    series: list[_Series],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    findings: list[Finding] = []
    for pair_index, (first, second) in enumerate(_candidate_series_pairs(series)):
        if checkpoint and pair_index % 128 == 0:
            checkpoint()
        length = len(first.cells)
        if length < MINIMUM_NEAR_DUPLICATE_LENGTH:
            continue
        first_values = first.values
        second_values = second.values
        equal = [
            settings.close(left, right)
            for left, right in zip(first_values, second_values, strict=True)
        ]
        mismatch_count = length - sum(equal)
        maximum_mismatches = max(1, math.floor(length * 0.2))
        if not 1 <= mismatch_count <= maximum_mismatches:
            continue
        if Counter(first_values) == Counter(second_values):
            continue
        similarity = (length - mismatch_count) / length
        locations = (first.location, second.location)
        risk = settings.risk_for_run(length - mismatch_count)
        findings.append(
            Finding(
                deterministic_finding_id(SERIES_NEAR_DUPLICATE_RULE_ID, locations),
                SERIES_NEAR_DUPLICATE_RULE_ID,
                FindingType.SUSPECTED_REUSE,
                risk,
                "数值序列仅有少量位置被修改",
                f"两列有 {length - mismatch_count}/{length} 个位置的完整值一致或近似一致。",
                locations,
                min(0.95, 0.7 + similarity * 0.25),
                {
                    "matched_count": length - mismatch_count,
                    "mismatch_count": mismatch_count,
                    "similarity": similarity,
                    "alignment": "按每列数值单元格出现顺序",
                    "first_series": _series_evidence(first),
                    "second_series": _series_evidence(second),
                    "paired_values": _paired_series(first, second, equal),
                },
            )
        )
        if len(findings) >= MAX_FINDINGS_PER_RULE:
            break
    return findings


def _find_robust_linear_series(
    series: list[_Series],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    findings: list[Finding] = []
    for pair_index, (first, second) in enumerate(_candidate_series_pairs(series)):
        if checkpoint and pair_index % 128 == 0:
            checkpoint()
        if len(first.cells) < 4:
            continue
        fitted = _theil_sen(first.values, second.values, settings)
        if fitted is None:
            continue
        slope, intercept, inliers, predictions = fitted
        if settings.close(slope, Decimal(1)) or settings.close(intercept, Decimal(0)):
            continue
        inlier_count = sum(inliers)
        if inlier_count < settings.medium_run_length or inlier_count / len(inliers) < 0.8:
            continue
        locations = (first.location, second.location)
        findings.append(
            Finding(
                deterministic_finding_id(SERIES_LINEAR_RULE_ID, locations),
                SERIES_LINEAR_RULE_ID,
                FindingType.SUSPECTED_REUSE,
                settings.risk_for_run(inlier_count),
                "数值序列符合稳健线性变换",
                (
                    "第二列约等于第一列乘以 "
                    f"{decimal_text(slope)} 再加 {decimal_text(intercept)}，"
                    f"{inlier_count}/{len(inliers)} 个位置符合。"
                ),
                locations,
                0.78,
                {
                    "slope": decimal_text(slope),
                    "intercept": decimal_text(intercept),
                    "matched_count": inlier_count,
                    "outlier_count": len(inliers) - inlier_count,
                    "alignment": "按每列数值单元格出现顺序；Theil-Sen 稳健拟合",
                    "first_series": _series_evidence(first),
                    "second_series": _series_evidence(second),
                    "paired_values": _paired_linear(first, second, predictions, inliers),
                },
            )
        )
        if len(findings) >= MAX_FINDINGS_PER_RULE:
            break
    return findings


def _theil_sen(
    first: tuple[Decimal, ...],
    second: tuple[Decimal, ...],
    settings: ExcelAnalysisSettings,
) -> tuple[Decimal, Decimal, list[bool], tuple[Decimal, ...]] | None:
    sample_count = min(len(first), 64)
    sampled_indexes = sorted(
        {
            round(position * (len(first) - 1) / max(1, sample_count - 1))
            for position in range(sample_count)
        }
    )
    slopes = [
        (second[right] - second[left]) / (first[right] - first[left])
        for left, right in combinations(sampled_indexes, 2)
        if first[right] != first[left]
    ]
    if not slopes:
        return None
    slope = _median(slopes)
    intercept = _median([right - slope * left for left, right in zip(first, second, strict=True)])
    predictions = tuple(slope * value + intercept for value in first)
    inliers = [
        settings.close(actual, predicted)
        for actual, predicted in zip(second, predictions, strict=True)
    ]
    return slope, intercept, inliers, predictions


def _find_exact_regions(
    cells: list[NumericCell],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    sheets: dict[tuple[str, str], dict[tuple[int, int], NumericCell]] = defaultdict(dict)
    for cell in cells:
        sheets[(cell.source_path, cell.sheet)][(cell.row, cell.column)] = cell
    signatures: dict[tuple[str, str, str, str], list[tuple[str, str, int, int]]] = defaultdict(list)
    for sheet_index, ((source, sheet), grid) in enumerate(sorted(sheets.items())):
        if checkpoint and sheet_index % 8 == 0:
            checkpoint()
        for row, column in sorted(grid):
            region = _region_values(grid, row, column, 2, 2)
            if region and len(set(region)) > 1 and set(region) != {"0", "1"}:
                signatures[tuple(region)].append((source, sheet, row, column))

    candidates: list[
        tuple[int, tuple[str, str, int, int], tuple[str, str, int, int], int, int]
    ] = []
    for positions in signatures.values():
        for first, second in combinations(positions[:MAX_GROUPS_PER_BUCKET], 2):
            if first[:2] == second[:2] and _rectangles_overlap(
                first[2], first[3], second[2], second[3], 2, 2
            ):
                continue
            height, width = _expand_equal_region(
                sheets[first[:2]], sheets[second[:2]], first[2:], second[2:]
            )
            candidates.append((height * width, first, second, height, width))
    candidates.sort(key=lambda item: (-item[0], item[1], item[2]))

    findings: list[Finding] = []
    covered: list[tuple[tuple[str, str, int, int], tuple[str, str, int, int], int, int]] = []
    for candidate_index, (_, first, second, height, width) in enumerate(candidates):
        if checkpoint and candidate_index % 64 == 0:
            checkpoint()
        if any(_candidate_is_covered(first, second, height, width, item) for item in covered):
            continue
        covered.append((first, second, height, width))
        first_location = _region_location(sheets[first[:2]], first, height, width)
        second_location = _region_location(sheets[second[:2]], second, height, width)
        locations = (first_location, second_location)
        paired = _paired_regions(
            sheets[first[:2]], sheets[second[:2]], first[2:], second[2:], height, width
        )
        area = height * width
        findings.append(
            Finding(
                deterministic_finding_id(REGION_EXACT_RULE_ID, locations),
                REGION_EXACT_RULE_ID,
                FindingType.EXACT_DUPLICATE,
                settings.risk_for_run(area),
                "连续二维数值区域完全重复",
                f"两个 {height} 行 × {width} 列的连续区域具有完全相同的完整数值。",
                locations,
                0.98,
                {
                    "matched_count": area,
                    "row_count": height,
                    "column_count": width,
                    "alignment": "按二维区域相对行列位置",
                    "first_series": _region_evidence(first_location, height, width),
                    "second_series": _region_evidence(second_location, height, width),
                    "paired_values": paired,
                },
            )
        )
        if len(findings) >= MAX_FINDINGS_PER_RULE:
            break
    return findings


def _region_values(
    grid: dict[tuple[int, int], NumericCell], row: int, column: int, height: int, width: int
) -> list[str] | None:
    region = [
        grid.get((row_offset, column_offset))
        for row_offset in range(row, row + height)
        for column_offset in range(column, column + width)
    ]
    return [cell.canonical_value for cell in region] if all(region) else None


def _expand_equal_region(
    first_grid: dict[tuple[int, int], NumericCell],
    second_grid: dict[tuple[int, int], NumericCell],
    first_start: tuple[int, int],
    second_start: tuple[int, int],
) -> tuple[int, int]:
    width_first_height = 2
    width_first_width = 2
    while _regions_equal(
        first_grid,
        second_grid,
        first_start,
        second_start,
        width_first_height,
        width_first_width + 1,
    ):
        width_first_width += 1
    while _regions_equal(
        first_grid,
        second_grid,
        first_start,
        second_start,
        width_first_height + 1,
        width_first_width,
    ):
        width_first_height += 1

    height_first_height = 2
    height_first_width = 2
    while _regions_equal(
        first_grid,
        second_grid,
        first_start,
        second_start,
        height_first_height + 1,
        height_first_width,
    ):
        height_first_height += 1
    while _regions_equal(
        first_grid,
        second_grid,
        first_start,
        second_start,
        height_first_height,
        height_first_width + 1,
    ):
        height_first_width += 1

    candidates = (
        (width_first_height, width_first_width),
        (height_first_height, height_first_width),
    )
    return max(candidates, key=lambda size: (size[0] * size[1], size[0], size[1]))


def _regions_equal(
    first_grid: dict[tuple[int, int], NumericCell],
    second_grid: dict[tuple[int, int], NumericCell],
    first_start: tuple[int, int],
    second_start: tuple[int, int],
    height: int,
    width: int,
) -> bool:
    first = _region_values(first_grid, *first_start, height, width)
    second = _region_values(second_grid, *second_start, height, width)
    return first is not None and first == second


def _rectangles_overlap(
    first_row: int,
    first_column: int,
    second_row: int,
    second_column: int,
    height: int,
    width: int,
) -> bool:
    return not (
        first_row + height <= second_row
        or second_row + height <= first_row
        or first_column + width <= second_column
        or second_column + width <= first_column
    )


def _candidate_is_covered(
    first: tuple[str, str, int, int],
    second: tuple[str, str, int, int],
    height: int,
    width: int,
    covered: tuple[tuple[str, str, int, int], tuple[str, str, int, int], int, int],
) -> bool:
    covered_first, covered_second, covered_height, covered_width = covered
    return _inside(first, height, width, covered_first, covered_height, covered_width) and _inside(
        second, height, width, covered_second, covered_height, covered_width
    )


def _inside(
    candidate: tuple[str, str, int, int],
    height: int,
    width: int,
    outer: tuple[str, str, int, int],
    outer_height: int,
    outer_width: int,
) -> bool:
    return (
        candidate[:2] == outer[:2]
        and outer[2] <= candidate[2]
        and outer[3] <= candidate[3]
        and candidate[2] + height <= outer[2] + outer_height
        and candidate[3] + width <= outer[3] + outer_width
    )


def _region_location(
    grid: dict[tuple[int, int], NumericCell],
    position: tuple[str, str, int, int],
    height: int,
    width: int,
) -> EvidenceLocation:
    source, sheet, row, column = position
    return EvidenceLocation(
        source,
        sheet,
        (
            f"{get_column_letter(column)}{row}:"
            f"{get_column_letter(column + width - 1)}{row + height - 1}"
        ),
        grid[(row, column)].hidden_sheet,
    )


def _paired_regions(
    first_grid: dict[tuple[int, int], NumericCell],
    second_grid: dict[tuple[int, int], NumericCell],
    first_start: tuple[int, int],
    second_start: tuple[int, int],
    height: int,
    width: int,
) -> list[dict[str, object]]:
    paired: list[dict[str, object]] = []
    for row_offset in range(height):
        for column_offset in range(width):
            first = first_grid[(first_start[0] + row_offset, first_start[1] + column_offset)]
            second = second_grid[(second_start[0] + row_offset, second_start[1] + column_offset)]
            paired.append(
                {
                    "position": f"{row_offset + 1},{column_offset + 1}",
                    "first_coordinate": first.coordinate,
                    "first_value": first.canonical_value,
                    "second_coordinate": second.coordinate,
                    "second_value": second.canonical_value,
                    "relation_result": "完整值相同",
                }
            )
    return paired


def _find_statistical_similarity(
    series: list[_Series],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None,
) -> list[Finding]:
    findings: list[Finding] = []
    for pair_index, (first, second) in enumerate(_candidate_series_pairs(series)):
        if checkpoint and pair_index % 128 == 0:
            checkpoint()
        if len(first.cells) < MINIMUM_STATISTICS_LENGTH:
            continue
        first_values = _finite_floats(first.values)
        second_values = _finite_floats(second.values)
        if first_values is None or second_values is None:
            continue
        if Counter(first.values) == Counter(second.values):
            continue
        if abs(_correlation(first_values, second_values)) >= 0.999:
            continue
        first_normalized = _normalized_sorted(first_values)
        second_normalized = _normalized_sorted(second_values)
        if first_normalized is None or second_normalized is None:
            continue
        correlation = _correlation(first_normalized, second_normalized)
        normalized_mae = sum(
            abs(left - right)
            for left, right in zip(first_normalized, second_normalized, strict=True)
        ) / len(first_normalized)
        first_mean, first_standard_deviation = _summary_statistics(first_values)
        second_mean, second_standard_deviation = _summary_statistics(second_values)
        summary_match = settings.close(
            Decimal(str(first_mean)), Decimal(str(second_mean))
        ) and settings.close(
            Decimal(str(first_standard_deviation)),
            Decimal(str(second_standard_deviation)),
        )
        distribution_match = correlation >= 0.999 and normalized_mae <= 0.03
        if not summary_match and not distribution_match:
            continue
        locations = (first.location, second.location)
        description = (
            "两列的均值和标准差异常一致，但完整数值并不相同；必须人工复核。"
            if summary_match
            else "两列的标准化排序分布高度相似；这只是统计线索，必须人工复核。"
        )
        findings.append(
            Finding(
                deterministic_finding_id(STATISTICS_RULE_ID, locations),
                STATISTICS_RULE_ID,
                FindingType.STATISTICAL_ANOMALY,
                RiskLevel.LOW,
                "数值序列统计特征异常相似",
                description,
                locations,
                0.45,
                {
                    "matched_count": len(first.cells),
                    "distribution_correlation": correlation,
                    "normalized_mae": normalized_mae,
                    "summary_match": summary_match,
                    "first_mean": first_mean,
                    "second_mean": second_mean,
                    "first_standard_deviation": first_standard_deviation,
                    "second_standard_deviation": second_standard_deviation,
                    "alignment": "标准化后按数值从小到大比较，仅作统计提示",
                    "first_series": _series_evidence(first),
                    "second_series": _series_evidence(second),
                    "paired_values": _paired_sorted(first, second),
                },
            )
        )
        if len(findings) >= MAX_FINDINGS_PER_RULE:
            break
    return findings


def _candidate_series_pairs(series: list[_Series]) -> list[tuple[_Series, _Series]]:
    by_length: dict[int, list[_Series]] = defaultdict(list)
    for item in series:
        by_length[len(item.cells)].append(item)
    pairs: list[tuple[_Series, _Series]] = []
    for bucket in by_length.values():
        if len(bucket) <= ALL_PAIRS_SERIES_LIMIT:
            pairs.extend(combinations(bucket, 2))
            continue
        fingerprints: dict[tuple[int, ...], list[_Series]] = defaultdict(list)
        for item in bucket:
            fingerprint = _coarse_fingerprint(item.values)
            if fingerprint is not None and len(fingerprints[fingerprint]) < MAX_GROUPS_PER_BUCKET:
                fingerprints[fingerprint].append(item)
        for candidates in fingerprints.values():
            pairs.extend(combinations(candidates, 2))
    return pairs


def _coarse_fingerprint(values: tuple[Decimal, ...]) -> tuple[int, ...] | None:
    converted = _finite_floats(values)
    if converted is None:
        return None
    mean = sum(converted) / len(converted)
    scale = max(abs(value - mean) for value in converted)
    if scale <= 1e-15:
        return None
    return tuple(round((value - mean) / scale * 100) for value in converted[:16])


def _finite_floats(values: tuple[Decimal, ...]) -> list[float] | None:
    try:
        converted = [float(value) for value in values]
    except (OverflowError, ValueError):
        return None
    return converted if all(math.isfinite(value) for value in converted) else None


def _normalized_sorted(values: list[float]) -> list[float] | None:
    mean, standard_deviation = _summary_statistics(values)
    variance = standard_deviation**2
    if variance <= 1e-30:
        return None
    return sorted((value - mean) / standard_deviation for value in values)


def _summary_statistics(values: list[float]) -> tuple[float, float]:
    mean = sum(values) / len(values)
    variance = sum((value - mean) ** 2 for value in values) / len(values)
    return mean, math.sqrt(max(0.0, variance))


def _correlation(first: list[float], second: list[float]) -> float:
    first_mean = sum(first) / len(first)
    second_mean = sum(second) / len(second)
    first_centered = [value - first_mean for value in first]
    second_centered = [value - second_mean for value in second]
    denominator = math.sqrt(
        sum(value * value for value in first_centered)
        * sum(value * value for value in second_centered)
    )
    if denominator <= 1e-30:
        return 0.0
    return (
        sum(left * right for left, right in zip(first_centered, second_centered, strict=True))
        / denominator
    )


def _median(values: list[Decimal]) -> Decimal:
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[midpoint]
    return (ordered[midpoint - 1] + ordered[midpoint]) / 2


def _series_evidence(series: _Series) -> dict[str, object]:
    return {
        "source_path": series.source_path,
        "sheet": series.sheet,
        "column": get_column_letter(series.column),
        "coordinates": [cell.coordinate for cell in series.cells],
        "values": [cell.canonical_value for cell in series.cells],
    }


def _series_range_location(series: _Series, cells: tuple[NumericCell, ...]) -> EvidenceLocation:
    column = get_column_letter(series.column)
    return EvidenceLocation(
        series.source_path,
        series.sheet,
        f"{column}{cells[0].row}:{column}{cells[-1].row}",
        cells[0].hidden_sheet,
    )


def _series_slice_evidence(series: _Series, cells: tuple[NumericCell, ...]) -> dict[str, object]:
    return {
        "source_path": series.source_path,
        "sheet": series.sheet,
        "column": get_column_letter(series.column),
        "coordinates": [cell.coordinate for cell in cells],
        "values": [cell.canonical_value for cell in cells],
    }


def _single_cell_evidence(cell: NumericCell) -> dict[str, object]:
    return {
        "source_path": cell.source_path,
        "sheet": cell.sheet,
        "column": get_column_letter(cell.column),
        "coordinates": [cell.coordinate],
        "values": [cell.canonical_value],
    }


def _region_evidence(location: EvidenceLocation, height: int, width: int) -> dict[str, object]:
    return {
        "source_path": location.source_path,
        "sheet": location.sheet or "",
        "coordinates": [location.coordinate or ""],
        "values": [],
        "row_count": height,
        "column_count": width,
    }


def _paired_series(first: _Series, second: _Series, matches: list[bool]) -> list[dict[str, object]]:
    return [
        {
            "position": index,
            "first_coordinate": left.coordinate,
            "first_value": left.canonical_value,
            "second_coordinate": right.coordinate,
            "second_value": right.canonical_value,
            "relation_result": "一致/近似" if matches[index - 1] else "不一致",
        }
        for index, (left, right) in enumerate(zip(first.cells, second.cells, strict=True), 1)
    ]


def _paired_linear(
    first: _Series,
    second: _Series,
    predictions: tuple[Decimal, ...],
    inliers: list[bool],
) -> list[dict[str, object]]:
    return [
        {
            "position": index,
            "first_coordinate": left.coordinate,
            "first_value": left.canonical_value,
            "second_coordinate": right.coordinate,
            "second_value": right.canonical_value,
            "relation_result": decimal_text(predicted),
            "matches": inliers[index - 1],
        }
        for index, (left, right, predicted) in enumerate(
            zip(first.cells, second.cells, predictions, strict=True), 1
        )
    ]


def _paired_sorted(first: _Series, second: _Series) -> list[dict[str, object]]:
    first_cells = sorted(first.cells, key=lambda cell: Decimal(cell.canonical_value))
    second_cells = sorted(second.cells, key=lambda cell: Decimal(cell.canonical_value))
    return [
        {
            "position": index,
            "first_coordinate": left.coordinate,
            "first_value": left.canonical_value,
            "second_coordinate": right.coordinate,
            "second_value": right.canonical_value,
            "relation_result": "标准化分布位置",
        }
        for index, (left, right) in enumerate(zip(first_cells, second_cells, strict=True), 1)
    ]

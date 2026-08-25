from __future__ import annotations

import math
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
from decimal import Decimal
from itertools import combinations

from openpyxl.utils import get_column_letter

from medical_image_check.domain.excel_settings import ExcelAnalysisSettings, decimal_text
from medical_image_check.domain.models import (
    EvidenceLocation,
    Finding,
    FindingType,
    RiskLevel,
    deterministic_finding_id,
)
from medical_image_check.engines.excel_patterns import find_excel_pattern_findings
from medical_image_check.infrastructure.spreadsheets import NumericCell

APPROXIMATE_RULE_ID = "excel.value.approximate"
SERIES_EXACT_RULE_ID = "excel.series.exact"
SERIES_SCALE_RULE_ID = "excel.series.scale"
SERIES_OFFSET_RULE_ID = "excel.series.offset"
SERIES_SUM_RULE_ID = "excel.series.target_sum"
SERIES_PRODUCT_RULE_ID = "excel.series.target_product"
SERIES_DIFFERENCE_RULE_ID = "excel.series.target_difference"
SERIES_QUOTIENT_RULE_ID = "excel.series.target_quotient"

APPROXIMATE_BANDS = (
    (Decimal("0.0001"), "0.01%"),
    (Decimal("0.001"), "0.1%"),
    (Decimal("0.01"), "1%"),
)
MINIMUM_SERIES_LENGTH = 3
ALL_PAIRS_SERIES_LIMIT = 250
MAX_APPROXIMATE_FINDINGS = 300


@dataclass(frozen=True, slots=True)
class _ValueGroup:
    canonical_value: str
    value: Decimal
    cells: tuple[NumericCell, ...]


@dataclass(frozen=True, slots=True)
class _NumericSeries:
    source_path: str
    sheet: str
    column: int
    cells: tuple[NumericCell, ...]

    @property
    def values(self) -> tuple[Decimal, ...]:
        return tuple(Decimal(cell.canonical_value) for cell in self.cells)

    @property
    def location(self) -> EvidenceLocation:
        column_name = get_column_letter(self.column)
        return EvidenceLocation(
            source_path=self.source_path,
            sheet=self.sheet,
            coordinate=f"{column_name}{self.cells[0].row}:{column_name}{self.cells[-1].row}",
            hidden_sheet=self.cells[0].hidden_sheet,
        )


def find_advanced_excel_findings(
    cells: list[NumericCell],
    settings: ExcelAnalysisSettings | None = None,
    checkpoint: Callable[[], None] | None = None,
) -> list[Finding]:
    configured = settings or ExcelAnalysisSettings()
    findings = _find_approximate_values(cells, configured)
    if checkpoint:
        checkpoint()
    findings.extend(_find_series_relations(cells, configured, checkpoint))
    if checkpoint:
        checkpoint()
    findings.extend(find_excel_pattern_findings(cells, configured, checkpoint))
    return findings


def _find_approximate_values(
    cells: list[NumericCell], settings: ExcelAnalysisSettings
) -> list[Finding]:
    grouped: dict[str, list[NumericCell]] = defaultdict(list)
    for cell in cells:
        grouped[cell.canonical_value].append(cell)
    value_groups = [
        _ValueGroup(canonical, Decimal(canonical), tuple(grouped_cells))
        for canonical, grouped_cells in sorted(grouped.items())
    ]

    approximate_groups: dict[tuple[int, ...], str] = {}
    positive = sorted(
        ((group.value, index) for index, group in enumerate(value_groups) if group.value > 0),
        key=lambda item: item[0],
    )
    negative = sorted(
        ((abs(group.value), index) for index, group in enumerate(value_groups) if group.value < 0),
        key=lambda item: item[0],
    )
    near_zero = sorted(
        (
            (group.value, index)
            for index, group in enumerate(value_groups)
            if abs(group.value) <= settings.absolute_tolerance
        ),
        key=lambda item: item[0],
    )
    for group_indexes in _anchored_clusters(
        near_zero,
        absolute_tolerance=settings.absolute_tolerance,
    ):
        approximate_groups[tuple(sorted(group_indexes))] = "绝对容差"

    for threshold, label in APPROXIMATE_BANDS:
        for ordered in (positive, negative):
            for group_indexes in _anchored_clusters(ordered, relative_tolerance=threshold):
                approximate_groups.setdefault(tuple(sorted(group_indexes)), label)
    custom_threshold = settings.custom_relative_tolerance
    if custom_threshold > 0 and custom_threshold not in {
        threshold for threshold, _ in APPROXIMATE_BANDS
    }:
        custom_label = f"自定义 {decimal_text(settings.custom_relative_tolerance_percent)}%"
        for ordered in (positive, negative):
            for group_indexes in _anchored_clusters(ordered, relative_tolerance=custom_threshold):
                approximate_groups.setdefault(tuple(sorted(group_indexes)), custom_label)

    findings: list[Finding] = []
    for group_indexes, band in sorted(approximate_groups.items()):
        groups = [value_groups[index] for index in group_indexes]
        ordered_values = sorted(group.value for group in groups)
        difference = ordered_values[-1] - ordered_values[0]
        relative_error = max(
            _relative_error(first.value, second.value) for first, second in combinations(groups, 2)
        )
        grouped_cells = sorted(
            (cell for group in groups for cell in group.cells),
            key=lambda cell: (cell.source_path, cell.sheet, cell.row, cell.column),
        )
        locations = tuple(cell.location for cell in grouped_cells)
        confidence = {
            "绝对容差": 0.7,
            "0.01%": 0.65,
            "0.1%": 0.55,
            "1%": 0.45,
        }.get(band, 0.4)
        findings.append(
            Finding(
                finding_id=deterministic_finding_id(APPROXIMATE_RULE_ID, locations),
                rule_id=APPROXIMATE_RULE_ID,
                finding_type=FindingType.HIGH_SIMILARITY,
                risk=RiskLevel.LOW,
                title="数值近似一致",
                description=(
                    f"{len(grouped_cells)} 个单元格中的 {len(groups)} 个不同完整数值"
                    f"落入 {band} 近似档位。"
                ),
                locations=locations,
                confidence=confidence,
                details={
                    "values": [group.canonical_value for group in groups],
                    "absolute_difference": str(difference),
                    "relative_error": float(relative_error),
                    "relative_error_percent": float(relative_error * 100),
                    "tolerance_band": band,
                    "cell_count": len(grouped_cells),
                    "distinct_value_count": len(groups),
                    "cells": _cell_evidence(tuple(grouped_cells)),
                },
            )
        )
    findings.sort(
        key=lambda item: (
            float(item.details["relative_error"]),
            -int(item.details["distinct_value_count"]),
            -int(item.details["cell_count"]),
            item.finding_id,
        )
    )
    return findings[:MAX_APPROXIMATE_FINDINGS]


def _anchored_clusters(
    ordered: list[tuple[Decimal, int]],
    *,
    relative_tolerance: Decimal | None = None,
    absolute_tolerance: Decimal | None = None,
) -> list[tuple[int, ...]]:
    clusters: list[tuple[int, ...]] = []
    position = 0
    while position < len(ordered):
        anchor_value = ordered[position][0]
        if relative_tolerance is not None:
            upper = anchor_value / (Decimal(1) - relative_tolerance)
        elif absolute_tolerance is not None:
            upper = anchor_value + absolute_tolerance
        else:
            raise ValueError("近似值分组必须提供相对或绝对容差")
        end = position + 1
        while end < len(ordered) and ordered[end][0] <= upper:
            end += 1
        if end - position >= 2:
            clusters.append(tuple(index for _, index in ordered[position:end]))
            position = end
        else:
            position += 1
    return clusters


def _relative_error(first: Decimal, second: Decimal) -> Decimal:
    scale = max(abs(first), abs(second))
    if scale == 0:
        return Decimal(0)
    return abs(first - second) / scale


def _find_series_relations(
    cells: list[NumericCell],
    settings: ExcelAnalysisSettings,
    checkpoint: Callable[[], None] | None = None,
) -> list[Finding]:
    series = _collect_series(cells, settings.medium_run_length)
    findings: list[Finding] = []
    for pair_index, (first_index, second_index) in enumerate(
        sorted(_series_candidate_pairs(series))
    ):
        if checkpoint and pair_index % 128 == 0:
            checkpoint()
        findings.extend(_compare_series(series[first_index], series[second_index], settings))
    return findings


def _collect_series(
    cells: list[NumericCell], minimum_length: int = MINIMUM_SERIES_LENGTH
) -> list[_NumericSeries]:
    grouped: dict[tuple[str, str, int], list[NumericCell]] = defaultdict(list)
    for cell in cells:
        grouped[(cell.source_path, cell.sheet, cell.column)].append(cell)
    return [
        _NumericSeries(
            source_path=key[0],
            sheet=key[1],
            column=key[2],
            cells=tuple(sorted(grouped_cells, key=lambda cell: cell.row)),
        )
        for key, grouped_cells in sorted(grouped.items())
        if len(grouped_cells) >= minimum_length
    ]


def _series_candidate_pairs(series: list[_NumericSeries]) -> set[tuple[int, int]]:
    by_length: dict[int, list[int]] = defaultdict(list)
    for index, item in enumerate(series):
        by_length[len(item.cells)].append(index)

    candidates: set[tuple[int, int]] = set()
    for indexes in by_length.values():
        if len(indexes) <= ALL_PAIRS_SERIES_LIMIT:
            candidates.update(combinations(indexes, 2))
            continue
        shape_index: dict[tuple[str, tuple[int, ...]], list[int]] = defaultdict(list)
        for index in indexes:
            values = series[index].values
            linear_shape = _series_shape(values)
            if linear_shape is not None:
                shape_index[("linear", linear_shape)].append(index)
            log_shape = _series_shape(values, logarithmic=True)
            if log_shape is not None:
                shape_index[("log", log_shape)].append(index)
        for grouped_indexes in shape_index.values():
            candidates.update(combinations(grouped_indexes, 2))
    return candidates


def _series_shape(
    values: tuple[Decimal, ...],
    *,
    logarithmic: bool = False,
) -> tuple[int, ...] | None:
    try:
        converted = [
            math.log(abs(float(value))) if logarithmic else float(value) for value in values
        ]
    except (OverflowError, ValueError):
        return None
    if not all(math.isfinite(value) for value in converted):
        return None
    if logarithmic and any(value == 0 for value in values):
        return None
    mean = sum(converted) / len(converted)
    centered = [value - mean for value in converted]
    scale = max(abs(value) for value in centered)
    if scale <= 1e-15:
        return None
    normalized = [round(value / scale * 10_000) for value in centered]
    first_nonzero = next((value for value in normalized if value), 1)
    if first_nonzero < 0:
        normalized = [-value for value in normalized]
    return tuple(normalized)


def _compare_series(
    first: _NumericSeries,
    second: _NumericSeries,
    settings: ExcelAnalysisSettings,
) -> list[Finding]:
    first_values = first.values
    second_values = second.values
    if len(first_values) != len(second_values):
        return []
    if not _is_informative_series(first_values) or not _is_informative_series(second_values):
        return []

    relations: list[tuple[str, str, str, Decimal, tuple[Decimal, ...]]] = []
    series_exact = all(
        first_value == second_value
        for first_value, second_value in zip(first_values, second_values, strict=True)
    )
    if series_exact:
        relations.append(
            (
                SERIES_EXACT_RULE_ID,
                "数值序列完全重复",
                "两列按数值出现顺序具有完全相同的完整值。",
                Decimal(1),
                second_values,
            )
        )
    else:
        factor = _constant_scale(first_values, second_values, settings)
        if factor is not None and not settings.close(factor, Decimal(1)):
            relations.append(
                (
                    SERIES_SCALE_RULE_ID,
                    "数值序列存在固定倍数",
                    f"第二列约等于第一列乘以固定倍数 {factor}。",
                    factor,
                    tuple(factor * value for value in first_values),
                )
            )
        offset = _constant_offset(first_values, second_values, settings)
        if offset is not None and not settings.close(offset, Decimal(0)):
            relations.append(
                (
                    SERIES_OFFSET_RULE_ID,
                    "数值序列存在固定偏移",
                    f"第二列约等于第一列加上固定值 {offset}。",
                    offset,
                    tuple(value + offset for value in first_values),
                )
            )

    summed = tuple(
        first_value + second_value
        for first_value, second_value in zip(first_values, second_values, strict=True)
    )
    sum_target = settings.target_for(summed)
    if sum_target is not None:
        relations.append(
            (
                SERIES_SUM_RULE_ID,
                "配对数值相加得到固定目标",
                f"两列配对相加持续得到整数目标 {sum_target}。",
                sum_target,
                summed,
            )
        )

    products = tuple(
        first_value * second_value
        for first_value, second_value in zip(first_values, second_values, strict=True)
    )
    product_target = settings.target_for(products)
    if product_target is not None and product_target != 0:
        relations.append(
            (
                SERIES_PRODUCT_RULE_ID,
                "配对数值相乘得到固定目标",
                f"两列配对相乘持续得到整数目标 {product_target}。",
                product_target,
                products,
            )
        )

    differences = tuple(
        first_value - second_value
        for first_value, second_value in zip(first_values, second_values, strict=True)
    )
    difference_target = settings.target_for(differences) if not series_exact else None
    if difference_target is not None:
        relations.append(
            (
                SERIES_DIFFERENCE_RULE_ID,
                "配对数值相减得到固定目标",
                f"第一列减第二列持续得到目标 {difference_target}。",
                difference_target,
                differences,
            )
        )

    if not series_exact and all(value != 0 for value in second_values):
        quotients = tuple(
            first_value / second_value
            for first_value, second_value in zip(first_values, second_values, strict=True)
        )
        quotient_target = settings.target_for(quotients)
        if quotient_target is not None:
            relations.append(
                (
                    SERIES_QUOTIENT_RULE_ID,
                    "配对数值相除得到固定目标",
                    f"第一列除以第二列持续得到目标 {quotient_target}。",
                    quotient_target,
                    quotients,
                )
            )

    return [_series_finding(first, second, *relation, settings) for relation in relations]


def _is_informative_series(values: tuple[Decimal, ...]) -> bool:
    """Reject constant and all-zero/identity sequences before relation mining.

    A varied sequence may legitimately contain zero or one, so those values are
    not removed individually. The gate only rejects columns that cannot carry a
    meaningful positional pattern.
    """

    return len(values) >= 2 and len(set(values)) >= 2 and any(value != 0 for value in values)


def _constant_scale(
    first: tuple[Decimal, ...],
    second: tuple[Decimal, ...],
    settings: ExcelAnalysisSettings,
) -> Decimal | None:
    ratios = [right / left for left, right in zip(first, second, strict=True) if left != 0]
    if not ratios:
        return None
    factor = _median(ratios)
    if all(settings.close(right, factor * left) for left, right in zip(first, second, strict=True)):
        return factor
    return None


def _constant_offset(
    first: tuple[Decimal, ...],
    second: tuple[Decimal, ...],
    settings: ExcelAnalysisSettings,
) -> Decimal | None:
    differences = [right - left for left, right in zip(first, second, strict=True)]
    offset = _median(differences)
    if all(settings.close(right, left + offset) for left, right in zip(first, second, strict=True)):
        return offset
    return None


def _median(values: list[Decimal]) -> Decimal:
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[midpoint]
    return (ordered[midpoint - 1] + ordered[midpoint]) / 2


def _series_finding(
    first: _NumericSeries,
    second: _NumericSeries,
    rule_id: str,
    title: str,
    description: str,
    parameter: Decimal,
    results: tuple[Decimal, ...],
    settings: ExcelAnalysisSettings,
) -> Finding:
    locations = (first.location, second.location)
    matched_count = len(first.cells)
    risk = settings.risk_for_run(matched_count)
    paired_values = [
        {
            "position": index,
            "first_coordinate": first_cell.coordinate,
            "first_value": first_cell.canonical_value,
            "second_coordinate": second_cell.coordinate,
            "second_value": second_cell.canonical_value,
            "relation_result": str(result),
        }
        for index, (first_cell, second_cell, result) in enumerate(
            zip(first.cells, second.cells, results, strict=True), start=1
        )
    ]
    return Finding(
        finding_id=deterministic_finding_id(rule_id, locations),
        rule_id=rule_id,
        finding_type=(
            FindingType.EXACT_DUPLICATE
            if rule_id == SERIES_EXACT_RULE_ID
            else FindingType.SUSPECTED_REUSE
        ),
        risk=risk,
        title=title,
        description=description,
        locations=locations,
        confidence=0.92 if risk == RiskLevel.HIGH else 0.82,
        details={
            "parameter": str(parameter),
            "matched_count": matched_count,
            "alignment": "按每列数值单元格出现顺序",
            "first_series": _series_evidence(first),
            "second_series": _series_evidence(second),
            "paired_values": paired_values,
        },
    )


def _series_evidence(series: _NumericSeries) -> dict[str, object]:
    return {
        "source_path": series.source_path,
        "sheet": series.sheet,
        "column": get_column_letter(series.column),
        "coordinates": [cell.coordinate for cell in series.cells],
        "values": [cell.canonical_value for cell in series.cells],
    }


def _cell_evidence(cells: tuple[NumericCell, ...]) -> list[dict[str, object]]:
    return [
        {
            "source_path": cell.source_path,
            "sheet": cell.sheet,
            "coordinate": cell.coordinate,
            "canonical_value": cell.canonical_value,
            "display_value": cell.display_value,
            "hidden_sheet": cell.hidden_sheet,
        }
        for cell in sorted(
            cells,
            key=lambda cell: (cell.source_path, cell.sheet, cell.row, cell.column),
        )
    ]

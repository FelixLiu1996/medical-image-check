from __future__ import annotations

import json
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from medical_image_check import __version__
from medical_image_check.domain.models import Finding, ReviewStatus, ScanResult
from medical_image_check.domain.project import Project
from medical_image_check.domain.review import marked_findings
from medical_image_check.services.report_common import REVIEW_LABELS, RISK_LABELS, TYPE_LABELS

FEEDBACK_SCHEMA_VERSION = 1
SUPPORTED_FEEDBACK_EXTENSIONS = frozenset({".xlsx", ".json"})


class FeedbackExporter:
    def export(
        self,
        result: ScanResult,
        destination: str | Path,
        project: Project | None = None,
    ) -> Path:
        findings = marked_findings(result)
        if not findings:
            raise ValueError("当前结果还没有人工反馈标记")
        output = Path(destination).expanduser().resolve()
        if output.suffix.lower() not in SUPPORTED_FEEDBACK_EXTENSIONS:
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)
        if output.suffix.lower() == ".json":
            return _export_json(result, findings, output, project)
        return _export_excel(result, findings, output, project)


def _export_json(
    result: ScanResult,
    findings: tuple[Finding, ...],
    output: Path,
    project: Project | None,
) -> Path:
    payload = _feedback_payload(result, findings, project)
    temporary = output.with_suffix(output.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    temporary.replace(output)
    return output


def _export_excel(
    result: ScanResult,
    findings: tuple[Finding, ...],
    output: Path,
    project: Project | None,
) -> Path:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "反馈概览"
    counts = Counter(finding.review_status for finding in findings)
    overview.append(["项目", "内容"])
    overview_rows = [
        ("反馈格式版本", FEEDBACK_SCHEMA_VERSION),
        ("软件版本", __version__),
        ("算法版本", result.algorithm_version),
        ("导出时间（UTC）", datetime.now(UTC).isoformat()),
        ("项目名称", project.name if project else "未关联项目"),
        ("项目 ID", project.project_id if project else ""),
        ("已标记结果", len(findings)),
        ("准确", counts[ReviewStatus.CONFIRMED]),
        ("误报", counts[ReviewStatus.FALSE_POSITIVE]),
        ("正常关联", counts[ReviewStatus.NORMAL]),
        ("原始文件", "未复制；清单仅包含定位信息和结构化证据"),
    ]
    for row in overview_rows:
        overview.append(row)
    _format_sheet(overview, (24, 86))

    worksheet = workbook.create_sheet("反馈清单")
    worksheet.append(
        [
            "结果 ID",
            "反馈",
            "风险",
            "类别",
            "规则",
            "标题",
            "说明",
            "置信度",
            "位置",
            "结构化位置 JSON",
            "结构化证据 JSON",
        ]
    )
    for finding in findings:
        worksheet.append(
            [
                finding.finding_id,
                REVIEW_LABELS[finding.review_status],
                RISK_LABELS[finding.risk],
                TYPE_LABELS[finding.finding_type],
                finding.rule_id,
                finding.title,
                finding.description,
                finding.confidence,
                "\n".join(location.display_text for location in finding.locations),
                json.dumps(
                    [_location_payload(location) for location in finding.locations],
                    ensure_ascii=False,
                    sort_keys=True,
                ),
                json.dumps(finding.details, ensure_ascii=False, sort_keys=True),
            ]
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _format_sheet(worksheet, (22, 12, 10, 14, 32, 26, 54, 12, 64, 64, 84))

    temporary = output.with_suffix(output.suffix + ".tmp")
    workbook.save(temporary)
    temporary.replace(output)
    return output


def _feedback_payload(
    result: ScanResult,
    findings: tuple[Finding, ...],
    project: Project | None,
) -> dict[str, object]:
    counts = Counter(finding.review_status for finding in findings)
    return {
        "schema_version": FEEDBACK_SCHEMA_VERSION,
        "software_version": __version__,
        "algorithm_version": result.algorithm_version,
        "generated_at": datetime.now(UTC).isoformat(),
        "project": ({"project_id": project.project_id, "name": project.name} if project else None),
        "summary": {
            "marked": len(findings),
            "confirmed": counts[ReviewStatus.CONFIRMED],
            "false_positive": counts[ReviewStatus.FALSE_POSITIVE],
            "normal": counts[ReviewStatus.NORMAL],
        },
        "raw_files_included": False,
        "feedback": [_finding_payload(finding) for finding in findings],
    }


def _finding_payload(finding: Finding) -> dict[str, object]:
    return {
        "finding_id": finding.finding_id,
        "review_status": finding.review_status.value,
        "rule_id": finding.rule_id,
        "risk": finding.risk.value,
        "finding_type": finding.finding_type.value,
        "title": finding.title,
        "description": finding.description,
        "confidence": finding.confidence,
        "locations": [_location_payload(location) for location in finding.locations],
        "evidence": finding.details,
    }


def _location_payload(location) -> dict[str, object]:
    return {
        "source_path": location.source_path,
        "sheet": location.sheet,
        "coordinate": location.coordinate,
        "hidden_sheet": location.hidden_sheet,
    }


def _format_sheet(worksheet, widths: tuple[int, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

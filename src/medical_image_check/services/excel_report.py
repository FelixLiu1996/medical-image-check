from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from medical_image_check import __version__
from medical_image_check.domain.image_settings import IMAGE_ANALYSIS_MODE_LABELS, ImageAnalysisMode
from medical_image_check.domain.models import RiskLevel, ScanResult
from medical_image_check.domain.project import Project
from medical_image_check.services.report_common import (
    CHANNEL_LABELS,
    EVIDENCE_KIND_LABELS,
    RELATIONSHIP_LABELS,
    REVIEW_LABELS,
    RISK_LABELS,
    TYPE_LABELS,
    attention_label,
)

REPORT_SCHEMA_VERSION = 1


def _panel_splitting_text(project: Project | None) -> str:
    if project is None or not project.panel_splitting_enabled:
        return "关闭"
    selected = sum(selection.selected for selection in project.panel_selections)
    return f"启用（已选 {selected}/{len(project.panel_selections)} 个子面板）"


class ExcelReportExporter:
    def export(
        self,
        result: ScanResult,
        destination: str | Path,
        project: Project | None = None,
    ) -> Path:
        output = Path(destination).expanduser().resolve()
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        overview = workbook.active
        overview.title = "扫描概览"
        self._write_overview(overview, result, project)
        self._write_findings(workbook, result)
        self._write_image_evidence(workbook, result)
        self._write_numeric_evidence(workbook, result)
        self._write_issues(workbook, result)
        self._write_sources(workbook, project)

        temporary = output.with_suffix(output.suffix + ".tmp")
        workbook.save(temporary)
        temporary.replace(output)
        return output

    @staticmethod
    def _write_overview(worksheet, result: ScanResult, project: Project | None) -> None:
        rows = [
            ("报告格式版本", REPORT_SCHEMA_VERSION),
            ("软件版本", __version__),
            ("算法版本", result.algorithm_version),
            ("扫描完成时间（UTC）", result.completed_at or "旧结果未记录"),
            ("生成时间（UTC）", datetime.now(UTC).isoformat()),
            ("项目名称", project.name if project else "未关联项目"),
            ("项目 ID", project.project_id if project else ""),
            ("连续数字片段最短位数", project.minimum_digit_run if project else ""),
            (
                "图片内容类型",
                IMAGE_ANALYSIS_MODE_LABELS[ImageAnalysisMode(project.image_analysis_mode)]
                if project
                else "",
            ),
            (
                "Western blot 单条带检测",
                "启用" if project and project.western_single_band_enabled else "关闭",
            ),
            (
                "复合图拆分",
                _panel_splitting_text(project),
            ),
            (
                "Excel 自定义相对容差",
                f"{project.excel_custom_relative_tolerance_percent}%" if project else "",
            ),
            (
                "Excel 绝对容差",
                project.excel_absolute_tolerance if project else "",
            ),
            (
                "Excel 运算目标",
                ", ".join(project.excel_operation_targets) if project else "",
            ),
            (
                "Excel 连续关系风险阈值",
                (
                    f"中风险 {project.excel_medium_run_length}；"
                    f"高风险 {project.excel_high_run_length}"
                )
                if project
                else "",
            ),
            ("扫描文件数", result.source_count),
            ("图片数", result.image_count),
            ("表格数", result.spreadsheet_count),
            ("结果数", len(result.findings)),
            ("高风险结果", sum(item.risk == RiskLevel.HIGH for item in result.findings)),
            ("中风险结果", sum(item.risk == RiskLevel.MEDIUM for item in result.findings)),
            ("低风险结果", sum(item.risk == RiskLevel.LOW for item in result.findings)),
            ("扫描提示数", len(result.issues)),
            ("用途说明", "本报告仅提供科研数据复核候选证据，不自动判定学术不端。"),
        ]
        worksheet.append(["项目", "内容"])
        for row in rows:
            worksheet.append(row)
        _format_sheet(worksheet, (24, 72))

    @staticmethod
    def _write_findings(workbook: Workbook, result: ScanResult) -> None:
        worksheet = workbook.create_sheet("查重结果")
        worksheet.append(
            [
                "结果 ID",
                "风险",
                "结果类别",
                "检测规则",
                "标题",
                "说明",
                "置信度",
                "位置 1",
                "位置 2",
                "其他位置",
                "证据详情",
                "人工复核",
                "候选层级",
            ]
        )
        for finding in result.findings:
            locations = [location.display_text for location in finding.locations]
            worksheet.append(
                [
                    finding.finding_id,
                    RISK_LABELS[finding.risk],
                    TYPE_LABELS[finding.finding_type],
                    finding.rule_id,
                    finding.title,
                    finding.description,
                    finding.confidence,
                    locations[0] if locations else "",
                    locations[1] if len(locations) > 1 else "",
                    "\n".join(locations[2:]),
                    _summary_details_json(finding.details),
                    REVIEW_LABELS[finding.review_status],
                    attention_label(finding.details),
                ]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _format_sheet(worksheet, (22, 8, 12, 24, 22, 48, 10, 48, 48, 48, 48, 12, 14))

    @staticmethod
    def _write_image_evidence(workbook: Workbook, result: ScanResult) -> None:
        worksheet = workbook.create_sheet("图像证据")
        worksheet.append(
            [
                "结果 ID",
                "检测规则",
                "来源 1",
                "页/面板 1",
                "区域 1",
                "条带数 1",
                "来源 2",
                "页/面板 2",
                "区域 2",
                "条带数 2",
                "匹配条带数",
                "条带结构相似度",
                "排列几何相似度",
                "背景纹理相似度",
                "掩膜重叠率",
                "变换",
                "单条带模式",
                "证据类型",
                "关系分类",
                "通道 1",
                "通道 2",
                "归一化互信息",
                "配准位移",
                "倍率 1",
                "倍率 2",
                "估算尺度比",
                "组织占比 1",
                "组织占比 2",
                "斑点数 1",
                "斑点数 2",
                "匹配斑点数",
                "斑点排列相似度",
                "斑点强度/形态轮廓相似度",
                "归一化排列误差",
                "斑点局部图像相似度",
                "最低单斑点相似度",
                "斑点缩放比",
                "斑点旋转角度",
                "斑点镜像",
            ]
        )
        for finding in result.findings:
            evidence_kind = finding.details.get("evidence_kind")
            if not (
                finding.rule_id.startswith("image.western_blot.")
                or evidence_kind
                in {
                    "western_blot",
                    "dot_blot",
                    "local_pattern",
                    "fluorescence",
                    "pathology",
                }
            ):
                continue
            details = finding.details
            locations = list(finding.locations)
            worksheet.append(
                [
                    finding.finding_id,
                    finding.rule_id,
                    locations[0].source_path if locations else "",
                    locations[0].coordinate if locations else "",
                    _region_text(details, "first"),
                    details.get("first_band_count", ""),
                    locations[1].source_path if len(locations) > 1 else "",
                    locations[1].coordinate if len(locations) > 1 else "",
                    _region_text(details, "second"),
                    details.get("second_band_count", ""),
                    details.get("matched_band_count", ""),
                    details.get("structure_similarity", ""),
                    details.get("geometry_similarity", ""),
                    details.get("background_similarity", ""),
                    details.get(
                        "band_mask_iou",
                        details.get(
                            "foreground_mask_iou",
                            details.get("tissue_mask_iou", ""),
                        ),
                    ),
                    details.get("transform_second_to_first", ""),
                    "是" if details.get("single_band_mode") else "否",
                    EVIDENCE_KIND_LABELS.get(
                        str(evidence_kind or "western_blot"),
                        str(evidence_kind or "western_blot"),
                    ),
                    RELATIONSHIP_LABELS.get(
                        str(details.get("relationship_class", "")),
                        str(details.get("relationship_class", "")),
                    ),
                    CHANNEL_LABELS.get(
                        str(details.get("first_channel", "")),
                        str(details.get("first_channel", "")),
                    ),
                    CHANNEL_LABELS.get(
                        str(details.get("second_channel", "")),
                        str(details.get("second_channel", "")),
                    ),
                    details.get("normalized_mutual_information", ""),
                    _alignment_shift_text(details),
                    details.get("first_magnification", ""),
                    details.get("second_magnification", ""),
                    details.get("estimated_scale_ratio", ""),
                    details.get("first_tissue_fraction", ""),
                    details.get("second_tissue_fraction", ""),
                    details.get("first_spot_count", ""),
                    details.get("second_spot_count", ""),
                    details.get("matched_spot_count", ""),
                    details.get("layout_similarity", ""),
                    details.get("profile_similarity", ""),
                    details.get("layout_error", ""),
                    details.get("appearance_similarity", ""),
                    details.get("minimum_spot_similarity", ""),
                    details.get("scale_second_to_first", ""),
                    details.get("rotation_degrees_second_to_first", ""),
                    "是" if details.get("mirrored") else "否",
                ]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _format_sheet(
            worksheet,
            (
                22,
                34,
                48,
                24,
                24,
                12,
                48,
                24,
                24,
                12,
                14,
                18,
                18,
                18,
                14,
                20,
                14,
                16,
                30,
                12,
                12,
                18,
                18,
                12,
                12,
                14,
                14,
                14,
                12,
                12,
                14,
                18,
                18,
                24,
                20,
                18,
                14,
                14,
                12,
            ),
        )

    @staticmethod
    def _write_numeric_evidence(workbook: Workbook, result: ScanResult) -> None:
        worksheet = workbook.create_sheet("数值证据")
        worksheet.append(
            [
                "结果 ID",
                "检测规则",
                "序号",
                "来源 1",
                "工作表 1",
                "坐标 1",
                "完整值 1",
                "读取值 1",
                "来源 2",
                "工作表 2",
                "坐标 2",
                "完整值 2",
                "关系结果",
                "隐藏工作表",
            ]
        )
        for finding in result.findings:
            cells = finding.details.get("cells")
            if isinstance(cells, list):
                for index, cell in enumerate(cells, start=1):
                    if not isinstance(cell, dict):
                        continue
                    worksheet.append(
                        [
                            finding.finding_id,
                            finding.rule_id,
                            index,
                            cell.get("source_path", ""),
                            cell.get("sheet", ""),
                            cell.get("coordinate", ""),
                            cell.get("canonical_value", ""),
                            cell.get("display_value", ""),
                            "",
                            "",
                            "",
                            "",
                            "",
                            "是" if cell.get("hidden_sheet") else "否",
                        ]
                    )
            paired_values = finding.details.get("paired_values")
            first_series = finding.details.get("first_series")
            second_series = finding.details.get("second_series")
            if not (
                isinstance(paired_values, list)
                and isinstance(first_series, dict)
                and isinstance(second_series, dict)
            ):
                continue
            for index, pair in enumerate(paired_values, start=1):
                if not isinstance(pair, dict):
                    continue
                worksheet.append(
                    [
                        finding.finding_id,
                        finding.rule_id,
                        index,
                        first_series.get("source_path", ""),
                        first_series.get("sheet", ""),
                        pair.get("first_coordinate", ""),
                        pair.get("first_value", ""),
                        pair.get("first_value", ""),
                        second_series.get("source_path", ""),
                        second_series.get("sheet", ""),
                        pair.get("second_coordinate", ""),
                        pair.get("second_value", ""),
                        pair.get("relation_result", ""),
                        "",
                    ]
                )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _format_sheet(
            worksheet,
            (22, 28, 8, 48, 20, 12, 22, 22, 48, 20, 12, 22, 22, 12),
        )

    @staticmethod
    def _write_issues(workbook: Workbook, result: ScanResult) -> None:
        worksheet = workbook.create_sheet("扫描提示")
        worksheet.append(["级别", "来源", "说明"])
        for issue in result.issues:
            worksheet.append([issue.severity, issue.source_path, issue.message])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _format_sheet(worksheet, (12, 60, 72))

    @staticmethod
    def _write_sources(workbook: Workbook, project: Project | None) -> None:
        worksheet = workbook.create_sheet("项目输入")
        worksheet.append(["序号", "原始路径"])
        for index, source in enumerate(project.source_paths if project else (), start=1):
            worksheet.append([index, source])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _format_sheet(worksheet, (10, 100))


def _format_sheet(worksheet, widths: tuple[int, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _summary_details_json(details: dict) -> str:
    summary = dict(details)
    cells = summary.get("cells")
    if isinstance(cells, list):
        summary["cells"] = f"共 {len(cells)} 行，详见“数值证据”工作表"
    paired_values = summary.get("paired_values")
    if isinstance(paired_values, list):
        summary["paired_values"] = f"共 {len(paired_values)} 行，详见“数值证据”工作表"
    for key in ("first_bands", "second_bands"):
        bands = summary.get(key)
        if isinstance(bands, list):
            summary[key] = f"共 {len(bands)} 条，详见原始结构化结果与“图像证据”工作表"
    encoded = json.dumps(summary, ensure_ascii=False, sort_keys=True)
    if len(encoded) <= 32_000:
        return encoded
    return encoded[:31_970] + "…（内容过长，请查看图像或数值证据工作表）"


def _region_text(details: dict, prefix: str) -> str:
    values = [details.get(f"{prefix}_region_{suffix}") for suffix in ("x", "y", "width", "height")]
    if any(value is None for value in values):
        return ""
    return ", ".join(str(value) for value in values)


def _alignment_shift_text(details: dict) -> str:
    horizontal = details.get("alignment_shift_x")
    vertical = details.get("alignment_shift_y")
    if horizontal is None and vertical is None:
        return ""
    return f"{horizontal or 0}, {vertical or 0}"

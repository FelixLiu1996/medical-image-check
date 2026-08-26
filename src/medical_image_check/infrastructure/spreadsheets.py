from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter, range_boundaries

from medical_image_check.domain.models import EvidenceLocation, ScanIssue


@dataclass(frozen=True, slots=True)
class NumericCell:
    source_path: str
    sheet: str
    row: int
    column: int
    coordinate: str
    canonical_value: str
    display_value: str
    hidden_sheet: bool = False
    column_header: str | None = None
    header_row: int | None = None
    formula: str | None = None

    @property
    def location(self) -> EvidenceLocation:
        return EvidenceLocation(
            source_path=self.source_path,
            sheet=self.sheet,
            coordinate=self.coordinate,
            hidden_sheet=self.hidden_sheet,
        )


@dataclass(frozen=True, slots=True)
class SpreadsheetReadResult:
    cells: tuple[NumericCell, ...]
    issues: tuple[ScanIssue, ...] = ()


@dataclass(frozen=True, slots=True)
class SpreadsheetPreview:
    source_path: str
    sheet: str
    start_row: int
    start_column: int
    values: tuple[tuple[str, ...], ...]
    target_min_row: int
    target_max_row: int
    target_min_column: int
    target_max_column: int

    def is_target(self, row: int, column: int) -> bool:
        return (
            self.target_min_row <= row <= self.target_max_row
            and self.target_min_column <= column <= self.target_max_column
        )


def read_spreadsheet_preview(
    location: EvidenceLocation,
    *,
    padding_rows: int = 1,
    padding_columns: int = 2,
    maximum_rows: int = 12,
    maximum_columns: int = 10,
) -> SpreadsheetPreview:
    """Read a bounded, display-only window around an evidence location."""

    target = _preview_target(location.coordinate)
    min_column, min_row, max_column, max_row = target
    start_row = max(1, min_row - padding_rows)
    start_column = max(1, min_column - padding_columns)
    end_row = min(max_row + padding_rows, start_row + maximum_rows - 1)
    end_column = min(max_column + padding_columns, start_column + maximum_columns - 1)
    if end_row - start_row + 1 > maximum_rows:
        end_row = start_row + maximum_rows - 1
    if end_column - start_column + 1 > maximum_columns:
        end_column = start_column + maximum_columns - 1

    path = Path(location.source_path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        sheet_name, values = _openpyxl_preview(
            path, location.sheet, start_row, end_row, start_column, end_column
        )
    elif suffix == ".xls":
        sheet_name, values = _xlrd_preview(
            path, location.sheet, start_row, end_row, start_column, end_column
        )
    elif suffix == ".csv":
        sheet_name, values = _csv_preview(path, start_row, end_row, start_column, end_column)
    else:
        raise ValueError(f"不支持预览的表格格式：{path.suffix or '无扩展名'}")
    return SpreadsheetPreview(
        source_path=str(path),
        sheet=sheet_name,
        start_row=start_row,
        start_column=start_column,
        values=values,
        target_min_row=min_row,
        target_max_row=max_row,
        target_min_column=min_column,
        target_max_column=max_column,
    )


def _preview_target(coordinate: str | None) -> tuple[int, int, int, int]:
    text = (coordinate or "").strip()
    row_match = re.fullmatch(r"第\s*(\d+)\s*行", text)
    if row_match:
        row = int(row_match.group(1))
        return 1, row, 10, row
    try:
        return range_boundaries(text)
    except ValueError as exc:
        raise ValueError(f"无法解析单元格位置：{text or '-'}") from exc


def _openpyxl_preview(
    path: Path,
    requested_sheet: str | None,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> tuple[str, tuple[tuple[str, ...], ...]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = (
            workbook[requested_sheet]
            if requested_sheet in workbook.sheetnames
            else workbook.worksheets[0]
        )
        rows = tuple(
            tuple(_preview_text(cell.value) for cell in row)
            for row in sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_column,
                max_col=end_column,
            )
        )
        return sheet.title, rows
    finally:
        workbook.close()


def _xlrd_preview(
    path: Path,
    requested_sheet: str | None,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> tuple[str, tuple[tuple[str, ...], ...]]:
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        sheet = (
            workbook.sheet_by_name(requested_sheet)
            if requested_sheet in workbook.sheet_names()
            else workbook.sheet_by_index(0)
        )
        rows = tuple(
            tuple(
                _preview_text(sheet.cell_value(row - 1, column - 1))
                if row <= sheet.nrows and column <= sheet.ncols
                else ""
                for column in range(start_column, end_column + 1)
            )
            for row in range(start_row, end_row + 1)
        )
        return sheet.name, rows
    finally:
        workbook.release_resources()


def _csv_preview(
    path: Path,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
) -> tuple[str, tuple[tuple[str, ...], ...]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeError:
            continue
    if text is None:
        raise ValueError("无法识别 CSV 编码")
    try:
        dialect = csv.Sniffer().sniff(text[:8192])
    except csv.Error:
        dialect = csv.excel
    selected: list[tuple[str, ...]] = []
    for row_index, row in enumerate(csv.reader(text.splitlines(), dialect), start=1):
        if row_index < start_row:
            continue
        if row_index > end_row:
            break
        selected.append(
            tuple(
                row[column - 1] if column <= len(row) else ""
                for column in range(start_column, end_column + 1)
            )
        )
    width = end_column - start_column + 1
    while len(selected) < end_row - start_row + 1:
        selected.append(tuple("" for _ in range(width)))
    return "CSV", tuple(selected)


def _preview_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def canonical_numeric(value: int | float | Decimal) -> str:
    if isinstance(value, bool):
        raise ValueError("布尔值不属于数值扫描范围")
    if isinstance(value, float) and not math.isfinite(value):
        raise ValueError("非有限数值不属于扫描范围")
    decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
    if not decimal_value.is_finite():
        raise ValueError("非有限数值不属于扫描范围")
    if decimal_value == 0:
        return "0"
    text = format(decimal_value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def canonical_digit_string(canonical_value: str) -> str:
    """Remove signs and separators while preserving every canonical numeric digit."""
    return "".join(character for character in canonical_value if character in "0123456789")


def _numeric_value(value: object, *, is_date: bool = False) -> tuple[str, str] | None:
    if is_date or isinstance(value, (bool, date, datetime, time)):
        return None
    if not isinstance(value, (int, float, Decimal)):
        return None
    try:
        return canonical_numeric(value), str(value)
    except ValueError:
        return None


class SpreadsheetReader:
    def read(self, path: str | Path) -> SpreadsheetReadResult:
        source = Path(path)
        suffix = source.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_openpyxl(source)
        if suffix == ".xls":
            return self._read_xlrd(source)
        if suffix == ".csv":
            return self._read_csv(source)
        raise ValueError(f"不支持的表格格式：{source.suffix or '无扩展名'}")

    def _read_openpyxl(self, path: Path) -> SpreadsheetReadResult:
        values_book = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        formulas_book = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        cells: list[NumericCell] = []
        issues: list[ScanIssue] = []
        try:
            for formulas_sheet in formulas_book.worksheets:
                values_sheet = values_book[formulas_sheet.title]
                hidden = formulas_sheet.sheet_state != "visible"
                missing_formula_results = 0
                headers: dict[int, tuple[str, int]] = {}
                for formula_row, value_row in zip_longest(
                    formulas_sheet.iter_rows(), values_sheet.iter_rows(), fillvalue=()
                ):
                    for formula_cell, value_cell in zip_longest(
                        formula_row, value_row, fillvalue=None
                    ):
                        if formula_cell is None or value_cell is None:
                            continue
                        if formula_cell.data_type == "f" and value_cell.value is None:
                            missing_formula_results += 1
                            continue
                        if formula_cell.data_type != "f" and isinstance(value_cell.value, str):
                            header = _header_text(value_cell.value)
                            if header:
                                headers[value_cell.column] = (header, value_cell.row)
                            continue
                        normalized = _numeric_value(
                            value_cell.value,
                            is_date=bool(value_cell.is_date or formula_cell.is_date),
                        )
                        if normalized is None:
                            continue
                        canonical, display = normalized
                        header = headers.get(value_cell.column)
                        cells.append(
                            NumericCell(
                                source_path=str(path),
                                sheet=formulas_sheet.title,
                                row=value_cell.row,
                                column=value_cell.column,
                                coordinate=value_cell.coordinate,
                                canonical_value=canonical,
                                display_value=display,
                                hidden_sheet=hidden,
                                column_header=header[0] if header else None,
                                header_row=header[1] if header else None,
                                formula=(
                                    str(formula_cell.value)
                                    if formula_cell.data_type == "f"
                                    else None
                                ),
                            )
                        )
                if missing_formula_results:
                    issues.append(
                        ScanIssue(
                            str(path),
                            f"工作表 {formulas_sheet.title} 有 {missing_formula_results} 个公式"
                            "没有保存计算结果，已跳过。",
                        )
                    )
        finally:
            formulas_book.close()
            values_book.close()
        return SpreadsheetReadResult(tuple(cells), tuple(issues))

    def _read_xlrd(self, path: Path) -> SpreadsheetReadResult:
        book = xlrd.open_workbook(path, on_demand=True)
        cells: list[NumericCell] = []
        try:
            for sheet in book.sheets():
                hidden = bool(getattr(sheet, "visibility", 0))
                headers: dict[int, tuple[str, int]] = {}
                for row_index in range(sheet.nrows):
                    for column_index in range(sheet.ncols):
                        cell = sheet.cell(row_index, column_index)
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            header = _header_text(cell.value)
                            if header:
                                headers[column_index + 1] = (header, row_index + 1)
                            continue
                        if cell.ctype != xlrd.XL_CELL_NUMBER:
                            continue
                        normalized = _numeric_value(cell.value)
                        if normalized is None:
                            continue
                        canonical, display = normalized
                        header = headers.get(column_index + 1)
                        cells.append(
                            NumericCell(
                                source_path=str(path),
                                sheet=sheet.name,
                                row=row_index + 1,
                                column=column_index + 1,
                                coordinate=f"{get_column_letter(column_index + 1)}{row_index + 1}",
                                canonical_value=canonical,
                                display_value=display,
                                hidden_sheet=hidden,
                                column_header=header[0] if header else None,
                                header_row=header[1] if header else None,
                            )
                        )
        finally:
            book.release_resources()
        return SpreadsheetReadResult(tuple(cells))

    def _read_csv(self, path: Path) -> SpreadsheetReadResult:
        text: str | None = None
        last_error: UnicodeError | None = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeError as exc:
                last_error = exc
        if text is None:
            raise ValueError(f"无法识别 CSV 编码：{last_error}")

        try:
            dialect = csv.Sniffer().sniff(text[:8192])
        except csv.Error:
            dialect = csv.excel

        cells: list[NumericCell] = []
        headers: dict[int, tuple[str, int]] = {}
        for row_index, row in enumerate(csv.reader(text.splitlines(), dialect), start=1):
            for column_index, raw_value in enumerate(row, start=1):
                normalized = self._parse_csv_number(raw_value)
                if normalized is None:
                    header = _header_text(raw_value)
                    if header:
                        headers[column_index] = (header, row_index)
                    continue
                canonical, display = normalized
                header = headers.get(column_index)
                cells.append(
                    NumericCell(
                        source_path=str(path),
                        sheet="CSV",
                        row=row_index,
                        column=column_index,
                        coordinate=f"{get_column_letter(column_index)}{row_index}",
                        canonical_value=canonical,
                        display_value=display,
                        column_header=header[0] if header else None,
                        header_row=header[1] if header else None,
                    )
                )
        return SpreadsheetReadResult(tuple(cells))

    @staticmethod
    def _parse_csv_number(raw_value: str) -> tuple[str, str] | None:
        stripped = raw_value.strip()
        if not stripped:
            return None
        is_percentage = stripped.endswith("%")
        candidate = stripped[:-1] if is_percentage else stripped
        candidate = candidate.replace(",", "")
        try:
            value = Decimal(candidate)
        except InvalidOperation:
            return None
        if is_percentage:
            value /= Decimal(100)
        try:
            return canonical_numeric(value), stripped
        except ValueError:
            return None


def _header_text(value: object) -> str | None:
    text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split())[:160]

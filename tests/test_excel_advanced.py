from pathlib import Path

import pytest
from openpyxl import Workbook

from medical_image_check.domain.excel_settings import ExcelAnalysisSettings
from medical_image_check.engines.excel_exact import ExactExcelDuplicateDetector


def _save_columns(path: Path, first: list[float], second: list[float] | None = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["第一组", "第二组"])
    row_count = max(len(first), len(second or []))
    for index in range(row_count):
        sheet.append(
            [
                first[index] if index < len(first) else None,
                second[index] if second is not None and index < len(second) else None,
            ]
        )
    workbook.save(path)


@pytest.mark.parametrize(
    ("values", "expected_band"),
    [
        ([100.0, 100.005], "0.01%"),
        ([100.0, 100.05], "0.1%"),
        ([100.0, 100.5], "1%"),
        ([0.0, 0.0000000000005], "绝对容差"),
    ],
)
def test_detector_finds_approximate_value_bands(
    tmp_path: Path,
    values: list[float],
    expected_band: str,
) -> None:
    path = tmp_path / f"approximate-{expected_band}.xlsx"
    _save_columns(path, values)

    findings, issues = ExactExcelDuplicateDetector().scan([path])

    assert issues == []
    approximate = [item for item in findings if item.rule_id == "excel.value.approximate"]
    assert len(approximate) == 1
    assert approximate[0].risk == "low"
    assert approximate[0].details["tolerance_band"] == expected_band
    assert len(approximate[0].details["cells"]) == 2


def test_detector_does_not_report_values_outside_one_percent(tmp_path: Path) -> None:
    path = tmp_path / "not-approximate.xlsx"
    _save_columns(path, [100.0, 102.0])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    assert not any(item.rule_id == "excel.value.approximate" for item in findings)


@pytest.mark.parametrize(
    ("first", "second", "rule_id", "parameter"),
    [
        ([1, 2, 3, 4], [2, 4, 6, 8], "excel.series.scale", "2"),
        ([1, 2, 3, 4], [6, 7, 8, 9], "excel.series.offset", "5"),
        ([1, 2, 3, 4], [99, 98, 97, 96], "excel.series.target_sum", "100"),
        ([1, 2, 4, 5], [100, 50, 25, 20], "excel.series.target_product", "100"),
        ([6, 7, 8, 9], [5, 6, 7, 8], "excel.series.target_difference", "1"),
        ([10, 20, 30, 40], [2, 4, 6, 8], "excel.series.target_quotient", "5"),
        ([1, 2, 3, 4], [1, 2, 3, 4], "excel.series.exact", "1"),
    ],
)
def test_detector_finds_repeated_series_relations(
    tmp_path: Path,
    first: list[float],
    second: list[float],
    rule_id: str,
    parameter: str,
) -> None:
    path = tmp_path / f"{rule_id}.xlsx"
    _save_columns(path, first, second)

    findings, issues = ExactExcelDuplicateDetector().scan([path])

    assert issues == []
    relations = [item for item in findings if item.rule_id == rule_id]
    assert len(relations) == 1
    assert relations[0].risk == "high"
    assert relations[0].details["parameter"] == parameter
    assert relations[0].details["matched_count"] == 4
    assert len(relations[0].details["paired_values"]) == 4


def test_three_repeated_series_relations_are_medium_risk(tmp_path: Path) -> None:
    path = tmp_path / "three-values.xlsx"
    _save_columns(path, [1, 2, 3], [10, 20, 30])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    scale = [item for item in findings if item.rule_id == "excel.series.scale"]
    assert len(scale) == 1
    assert scale[0].risk == "medium"


def test_custom_relative_tolerance_can_be_enabled(tmp_path: Path) -> None:
    path = tmp_path / "custom-tolerance.xlsx"
    _save_columns(path, [100, 101.5])
    settings = ExcelAnalysisSettings.from_values(2)

    findings, _ = ExactExcelDuplicateDetector(analysis_settings=settings).scan([path])

    approximate = [item for item in findings if item.rule_id == "excel.value.approximate"]
    assert len(approximate) == 1
    assert approximate[0].details["tolerance_band"] == "自定义 2%"


def test_detector_finds_two_cell_target_operation(tmp_path: Path) -> None:
    path = tmp_path / "cell-operation.xlsx"
    _save_columns(path, [2.5, 7.5])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    operations = [item for item in findings if item.rule_id == "excel.cell.target_operation"]
    assert any(
        item.details["operation"] == "add" and item.details["parameter"] == "10"
        for item in operations
    )
    assert all(item.risk == "low" for item in operations)


def test_cell_operation_uses_two_distinct_cells_with_the_same_value(tmp_path: Path) -> None:
    path = tmp_path / "same-value-operation.xlsx"
    _save_columns(path, [5, 5])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    operation = next(
        item
        for item in findings
        if item.rule_id == "excel.cell.target_operation" and item.details["operation"] == "add"
    )
    assert operation.details["parameter"] == "10"
    pair = operation.details["paired_values"][0]
    assert pair["first_coordinate"] != pair["second_coordinate"]


def test_detector_finds_shuffled_and_near_duplicate_series(tmp_path: Path) -> None:
    shuffled_path = tmp_path / "shuffled.xlsx"
    near_path = tmp_path / "near.xlsx"
    _save_columns(shuffled_path, [1, 2, 3, 4], [3, 1, 4, 2])
    _save_columns(near_path, [11, 22, 33, 44, 55], [11, 22, 33, 44, 99])

    shuffled, _ = ExactExcelDuplicateDetector().scan([shuffled_path])
    near, _ = ExactExcelDuplicateDetector().scan([near_path])

    shuffled_finding = next(item for item in shuffled if item.rule_id == "excel.series.shuffled")
    near_finding = next(item for item in near if item.rule_id == "excel.series.near_duplicate")
    assert shuffled_finding.details["matched_count"] == 4
    assert shuffled_finding.details["order_changed_count"] == 4
    assert near_finding.details["matched_count"] == 4
    assert near_finding.details["mismatch_count"] == 1


def test_detector_finds_exact_continuous_series_fragment(tmp_path: Path) -> None:
    path = tmp_path / "fragment.xlsx"
    _save_columns(path, [9, 11, 22, 33, 8], [7, 11, 22, 33, 6])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    fragment = next(item for item in findings if item.rule_id == "excel.series.fragment_exact")
    assert fragment.details["matched_count"] == 3
    assert fragment.risk == "medium"
    assert len(fragment.details["paired_values"]) == 3


def test_detector_finds_robust_linear_relation_with_outlier(tmp_path: Path) -> None:
    path = tmp_path / "robust-linear.xlsx"
    _save_columns(path, [1, 2, 3, 4, 5], [5, 8, 11, 14, 100])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    linear = next(item for item in findings if item.rule_id == "excel.series.linear")
    assert linear.details["slope"] == "3"
    assert linear.details["intercept"] == "2"
    assert linear.details["matched_count"] == 4
    assert linear.details["outlier_count"] == 1


def test_custom_tolerance_finds_whole_near_duplicate_series(tmp_path: Path) -> None:
    path = tmp_path / "near-series.xlsx"
    _save_columns(path, [100, 200, 300, 400, 500], [100.1, 200.2, 300.3, 400.4, 500.5])
    settings = ExcelAnalysisSettings.from_values(0.2)

    findings, _ = ExactExcelDuplicateDetector(analysis_settings=settings).scan([path])

    near = next(item for item in findings if item.rule_id == "excel.series.near_duplicate")
    assert near.details["matched_count"] == 5
    assert near.details["mismatch_count"] == 5
    assert near.details["out_of_tolerance_count"] == 0


def test_detector_finds_exact_two_dimensional_region(tmp_path: Path) -> None:
    path = tmp_path / "regions.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "第一块"
    second = workbook.create_sheet("第二块")
    matrix = [[12, 23, 34], [45, 56, 67], [78, 89, 91]]
    for row in matrix:
        first.append(row)
        second.append(row)
    workbook.save(path)

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    region = next(item for item in findings if item.rule_id == "excel.region.exact")
    assert region.details["row_count"] == 3
    assert region.details["column_count"] == 3
    assert region.details["matched_count"] == 9
    assert len(region.details["paired_values"]) == 9


def test_detector_marks_distribution_similarity_as_low_risk(tmp_path: Path) -> None:
    path = tmp_path / "statistics.xlsx"
    first = list(range(1, 11))
    second = [25, 1, 28, 4, 22, 7, 19, 10, 16, 13]
    _save_columns(path, first, second)

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    statistics = next(item for item in findings if item.rule_id == "excel.series.statistics")
    assert statistics.risk == "low"
    assert statistics.finding_type == "statistical_anomaly"
    assert statistics.details["distribution_correlation"] > 0.999


def test_detector_marks_matching_mean_and_standard_deviation(tmp_path: Path) -> None:
    path = tmp_path / "matching-summary.xlsx"
    _save_columns(path, [-1, -1, -1, -1, 1, 1, 1, 1], [-2, 0, 0, 0, 0, 0, 0, 2])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    statistics = next(item for item in findings if item.rule_id == "excel.series.statistics")
    assert statistics.risk == "low"
    assert statistics.details["summary_match"] is True
    assert statistics.details["first_mean"] == 0
    assert statistics.details["first_standard_deviation"] == 1

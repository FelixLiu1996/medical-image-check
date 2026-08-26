import json
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from medical_image_check.domain.models import (
    EvidenceLocation,
    Finding,
    FindingType,
    ReviewStatus,
    RiskLevel,
    ScanResult,
)
from medical_image_check.domain.project import Project
from medical_image_check.domain.review import (
    carry_forward_review_status,
    update_finding_review_status,
)
from medical_image_check.services.feedback_export import FeedbackExporter


def _finding(finding_id: str, status: ReviewStatus = ReviewStatus.PENDING) -> Finding:
    return Finding(
        finding_id=finding_id,
        rule_id="excel.series.target_sum",
        finding_type=FindingType.SUSPECTED_REUSE,
        risk=RiskLevel.MEDIUM,
        title="配对数值相加得到固定目标",
        description="连续配对值相加得到 2。",
        locations=(
            EvidenceLocation("source.xlsx", "Sheet1", "C3:C6"),
            EvidenceLocation("source.xlsx", "Sheet1", "D3:D6"),
        ),
        confidence=0.91,
        details={"target": "2", "matched_count": 4},
        review_status=status,
    )


def test_review_status_updates_and_carries_forward_by_stable_id() -> None:
    original = ScanResult(1, 0, 1, (_finding("stable"), _finding("other")))
    reviewed = update_finding_review_status(original, "stable", ReviewStatus.CONFIRMED)
    rescanned = ScanResult(1, 0, 1, (_finding("stable"), _finding("new")))

    carried = carry_forward_review_status(reviewed, rescanned)

    assert original.findings[0].review_status == ReviewStatus.PENDING
    assert reviewed.findings[0].review_status == ReviewStatus.CONFIRMED
    assert carried.findings[0].review_status == ReviewStatus.CONFIRMED
    assert carried.findings[1].review_status == ReviewStatus.PENDING

    upgraded = replace(rescanned, algorithm_version="test-2")
    not_carried = carry_forward_review_status(reviewed, upgraded)
    assert not_carried.findings[0].review_status == ReviewStatus.PENDING


def test_feedback_export_writes_only_marked_results_to_excel_and_json(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"unchanged")
    marked = _finding("marked", ReviewStatus.FALSE_POSITIVE)
    marked = replace(
        marked,
        locations=(
            EvidenceLocation(str(source), "Sheet1", "C3:C6"),
            EvidenceLocation(str(source), "Sheet1", "D3:D6"),
        ),
    )
    result = ScanResult(1, 0, 1, (marked, _finding("pending")), algorithm_version="test-1")
    project = Project.create("反馈测试").with_sources([source]).with_scan_result(result)
    exporter = FeedbackExporter()

    excel = exporter.export(result, tmp_path / "feedback", project)
    json_path = exporter.export(result, tmp_path / "feedback.json", project)

    workbook = load_workbook(excel, read_only=True)
    assert workbook.sheetnames == ["反馈概览", "反馈清单"]
    assert workbook["反馈清单"]["A2"].value == "marked"
    assert workbook["反馈清单"]["B2"].value == "误报"
    assert workbook["反馈清单"].max_row == 2
    workbook.close()
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["algorithm_version"] == "test-1"
    assert payload["raw_files_included"] is False
    assert payload["summary"] == {
        "marked": 1,
        "confirmed": 0,
        "false_positive": 1,
        "normal": 0,
    }
    assert [item["finding_id"] for item in payload["feedback"]] == ["marked"]
    assert source.read_bytes() == b"unchanged"


def test_feedback_export_requires_at_least_one_mark(tmp_path: Path) -> None:
    result = ScanResult(1, 0, 1, (_finding("pending"),))

    with pytest.raises(ValueError, match="没有人工反馈"):
        FeedbackExporter().export(result, tmp_path / "feedback.xlsx")

from datetime import datetime
from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook

from medical_image_check.engines.excel_exact import ExactExcelDuplicateDetector
from medical_image_check.infrastructure.spreadsheets import (
    SpreadsheetReader,
    canonical_digit_string,
)


def _save_workbook(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "实验A"
    first.append(["样本", "数值1", "数值2", "日期"])
    first.append(["a", 2.5, 3.5, "2026-08-25"])
    first.append(["b", 0, 1, None])

    second = workbook.create_sheet("实验B")
    second.sheet_state = "hidden"
    second.append(["样本", "数值1", "数值2"])
    second.append(["c", 2.5, 3.5])
    second.append(["d", 0, 1])
    workbook.save(path)


def _save_values(path: Path, values: list[int | float]) -> None:
    workbook = Workbook()
    for value in values:
        workbook.active.append([value])
    workbook.save(path)


def test_excel_detector_scans_hidden_sheets_and_exact_rows(tmp_path: Path) -> None:
    path = tmp_path / "experiment.xlsx"
    _save_workbook(path)

    findings, issues = ExactExcelDuplicateDetector().scan([path])

    assert issues == []
    assert any(finding.rule_id == "excel.row.exact" for finding in findings)
    value_findings = [finding for finding in findings if finding.rule_id == "excel.value.exact"]
    assert {finding.details["value"] for finding in value_findings} == {"2.5", "3.5"}
    assert any(location.hidden_sheet for finding in findings for location in finding.locations)


def test_reader_reports_formula_without_cached_result(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    workbook.save(path)

    result = SpreadsheetReader().read(path)

    assert result.cells == ()
    assert len(result.issues) == 1
    assert "没有保存计算结果" in result.issues[0].message


def test_csv_percentages_use_underlying_numeric_value(tmp_path: Path) -> None:
    path = tmp_path / "values.csv"
    path.write_text("组,数值\nA,50%\nB,0.5\n", encoding="utf-8")

    result = SpreadsheetReader().read(path)

    assert [cell.canonical_value for cell in result.cells] == ["0.5", "0.5"]


def test_xls_reader_ignores_dates_and_reads_numbers(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("原始数据")
    sheet.write(0, 0, 2.5)
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    sheet.write(0, 1, datetime(2026, 8, 25), date_style)
    workbook.save(str(path))

    result = SpreadsheetReader().read(path)

    assert [(cell.coordinate, cell.canonical_value) for cell in result.cells] == [("A1", "2.5")]


def test_canonical_digit_string_uses_complete_value_without_separators() -> None:
    assert canonical_digit_string("-120.034") == "120034"


@pytest.mark.parametrize(
    ("values", "risk", "fragment"),
    [
        ([14617, 94617], "low", "4617"),
        ([1123456, 9123457], "medium", "12345"),
        ([1123456789, 9123456780], "high", "12345678"),
    ],
)
def test_detector_ranks_shared_digit_fragments_by_length(
    tmp_path: Path,
    values: list[int],
    risk: str,
    fragment: str,
) -> None:
    path = tmp_path / f"fragment-{risk}.xlsx"
    _save_values(path, values)

    findings, issues = ExactExcelDuplicateDetector().scan([path])

    assert issues == []
    fragments = [item for item in findings if item.rule_id == "excel.digit_fragment"]
    assert len(fragments) == 1
    assert fragments[0].risk == risk
    assert fragment in fragments[0].details["fragments"]
    assert fragments[0].details["maximum_length"] == len(fragment)
    assert [cell["canonical_value"] for cell in fragments[0].details["cells"]] == [
        str(value) for value in values
    ]


def test_digit_fragment_ignores_internal_and_exact_value_repetition(tmp_path: Path) -> None:
    path = tmp_path / "internal-and-exact.xlsx"
    _save_values(path, [12341234, 12341234])

    findings, issues = ExactExcelDuplicateDetector().scan([path])

    assert issues == []
    assert any(item.rule_id == "excel.value.exact" for item in findings)
    assert not any(item.rule_id == "excel.digit_fragment" for item in findings)


def test_digit_fragment_ignores_sign_and_decimal_point(tmp_path: Path) -> None:
    path = tmp_path / "normalized-fragment.xlsx"
    _save_values(path, [-1234, 12.34])

    findings, _ = ExactExcelDuplicateDetector().scan([path])

    fragments = [item for item in findings if item.rule_id == "excel.digit_fragment"]
    assert len(fragments) == 1
    assert fragments[0].details["fragments"] == ["1234"]


def test_digit_fragment_minimum_length_is_configurable(tmp_path: Path) -> None:
    path = tmp_path / "configured-fragment.xlsx"
    _save_values(path, [14617, 94617])

    findings, _ = ExactExcelDuplicateDetector(minimum_digit_run=5).scan([path])

    assert not any(item.rule_id == "excel.digit_fragment" for item in findings)
    with pytest.raises(ValueError, match="3 到 12"):
        ExactExcelDuplicateDetector(minimum_digit_run=2)

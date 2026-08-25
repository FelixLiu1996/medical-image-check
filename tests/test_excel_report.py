import json
from pathlib import Path

from openpyxl import load_workbook

from medical_image_check.domain.models import (
    EvidenceLocation,
    Finding,
    FindingType,
    RiskLevel,
    ScanIssue,
    ScanResult,
)
from medical_image_check.domain.project import Project
from medical_image_check.services.excel_report import ExcelReportExporter


def test_excel_report_contains_overview_findings_issues_and_sources(tmp_path: Path) -> None:
    source = tmp_path / "data.xlsx"
    source.write_bytes(b"unchanged-source")
    locations = (
        EvidenceLocation(str(source), "实验A", "B2"),
        EvidenceLocation(str(source), "实验B", "C4", True),
    )
    finding = Finding(
        finding_id="finding-1",
        rule_id="excel.value.exact",
        finding_type=FindingType.EXACT_DUPLICATE,
        risk=RiskLevel.LOW,
        title="数值完全相同",
        description="完整数值 2.5 重复。",
        locations=locations,
        details={
            "value": "2.5",
            "count": 2,
            "cells": [{"coordinate": "B2", "canonical_value": "2.5"}],
        },
    )
    result = ScanResult(
        source_count=1,
        image_count=0,
        spreadsheet_count=1,
        findings=(finding,),
        issues=(ScanIssue(str(source), "公式无缓存"),),
    )
    project = Project.create("报告测试").with_sources([source]).with_scan_result(result)

    output = ExcelReportExporter().export(result, tmp_path / "report", project)

    assert output.name == "report.xlsx"
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "扫描概览",
        "查重结果",
        "图像证据",
        "数值证据",
        "扫描提示",
        "项目输入",
    ]
    findings = workbook["查重结果"]
    assert findings["A2"].value == "finding-1"
    assert findings["B2"].value == "低"
    assert "实验A" in findings["H2"].value
    assert "实验B" in findings["I2"].value
    evidence = json.loads(findings["K2"].value)
    assert "详见“数值证据”" in evidence["cells"]
    numeric_evidence = workbook["数值证据"]
    assert numeric_evidence["F2"].value == "B2"
    assert numeric_evidence["G2"].value == "2.5"
    assert workbook["扫描提示"]["C2"].value == "公式无缓存"
    assert workbook["项目输入"]["B2"].value == str(source.resolve())
    overview = {
        row[0].value: row[1].value for row in workbook["扫描概览"].iter_rows(min_row=2, max_col=2)
    }
    assert overview["Excel 自定义相对容差"] == "0.0%"
    assert overview["Excel 运算目标"] == "0, 1, 10, 100, 1000"
    assert overview["Excel 连续关系风险阈值"] == "中风险 3；高风险 4"
    workbook.close()
    assert source.read_bytes() == b"unchanged-source"


def test_excel_report_contains_structured_western_blot_evidence(tmp_path: Path) -> None:
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    locations = (EvidenceLocation(str(first)), EvidenceLocation(str(second)))
    finding = Finding(
        finding_id="western-1",
        rule_id="image.western_blot.panel_reuse",
        finding_type=FindingType.SUSPECTED_REUSE,
        risk=RiskLevel.MEDIUM,
        title="Western blot 面板或泳道疑似复用",
        description="多个证据共同匹配。",
        locations=locations,
        details={
            "first_region_x": 10,
            "first_region_y": 20,
            "first_region_width": 100,
            "first_region_height": 40,
            "second_region_x": 30,
            "second_region_y": 40,
            "second_region_width": 120,
            "second_region_height": 50,
            "first_band_count": 4,
            "second_band_count": 4,
            "matched_band_count": 4,
            "structure_similarity": 0.95,
            "geometry_similarity": 0.91,
            "background_similarity": 0.87,
            "band_mask_iou": 0.8,
            "transform_second_to_first": "flip_horizontal",
            "single_band_mode": False,
            "first_bands": [{"x": 10, "y": 20, "width": 20, "height": 5}],
            "second_bands": [{"x": 30, "y": 40, "width": 24, "height": 6}],
        },
    )
    result = ScanResult(2, 2, 0, (finding,))

    output = ExcelReportExporter().export(result, tmp_path / "western-report.xlsx")
    workbook = load_workbook(output, read_only=True)

    image_evidence = workbook["图像证据"]
    assert image_evidence["A2"].value == "western-1"
    assert image_evidence["E2"].value == "10, 20, 100, 40"
    assert image_evidence["K2"].value == 4
    assert image_evidence["N2"].value == 0.87
    assert image_evidence["Q2"].value == "否"
    details = json.loads(workbook["查重结果"]["K2"].value)
    assert "图像证据" in details["first_bands"]
    workbook.close()


def test_excel_report_contains_fluorescence_and_pathology_evidence(tmp_path: Path) -> None:
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    locations = (EvidenceLocation(str(first)), EvidenceLocation(str(second)))
    fluorescence = Finding(
        finding_id="fluorescence-1",
        rule_id="image.fluorescence.merge_component",
        finding_type=FindingType.NORMAL_RELATION,
        risk=RiskLevel.LOW,
        title="荧光单通道与 Merge 成分对应",
        description="正常关系。",
        locations=locations,
        details={
            "evidence_kind": "fluorescence",
            "relationship_class": "normal_merge_component",
            "first_channel": "blue",
            "second_channel": "blue",
            "structure_similarity": 0.96,
            "foreground_mask_iou": 0.82,
            "normalized_mutual_information": 0.75,
            "alignment_shift_x": 1.5,
            "alignment_shift_y": -2.0,
        },
    )
    pathology = Finding(
        finding_id="pathology-1",
        rule_id="image.pathology.same_region_different_magnification",
        finding_type=FindingType.NORMAL_RELATION,
        risk=RiskLevel.LOW,
        title="病理图疑似同一区域的不同倍率",
        description="正常关系。",
        locations=locations,
        details={
            "evidence_kind": "pathology",
            "relationship_class": "normal_different_magnification",
            "structure_similarity": 0.94,
            "tissue_mask_iou": 0.8,
            "first_magnification": 10.0,
            "second_magnification": 40.0,
            "estimated_scale_ratio": 4.0,
            "first_tissue_fraction": 0.65,
            "second_tissue_fraction": 0.72,
        },
    )
    result = ScanResult(2, 2, 0, (fluorescence, pathology))

    output = ExcelReportExporter().export(result, tmp_path / "specialized-report.xlsx")
    workbook = load_workbook(output, read_only=True)
    evidence = workbook["图像证据"]

    assert evidence["R2"].value == "荧光"
    assert evidence["T2"].value == "蓝色通道"
    assert evidence["V2"].value == 0.75
    assert evidence["W2"].value == "1.5, -2.0"
    assert evidence["R3"].value == "病理"
    assert evidence["X3"].value == 10.0
    assert evidence["Y3"].value == 40.0
    assert evidence["Z3"].value == 4.0
    workbook.close()

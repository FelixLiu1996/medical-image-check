from pathlib import Path
from threading import Event, Thread

import cv2
import numpy as np
import pytest
from openpyxl import Workbook

from medical_image_check.domain.panels import PanelSelection
from medical_image_check.services.basic_scan import (
    BasicScanService,
    ScanCancelled,
    ScanControl,
    ScanMode,
)


def test_basic_scan_collects_directory_and_reports_duplicates(tmp_path: Path) -> None:
    image = np.arange(32 * 32, dtype=np.uint8).reshape(32, 32)
    first_image = tmp_path / "one.png"
    assert cv2.imwrite(str(first_image), image)
    (tmp_path / "two.png").write_bytes(first_image.read_bytes())
    (tmp_path / "ignored.txt").write_text("same", encoding="utf-8")

    workbook = Workbook()
    workbook.active.append([2.5, 3.5])
    workbook.active.append([2.5, 3.5])
    workbook.save(tmp_path / "data.xlsx")

    progress: list[tuple[int, int, str]] = []
    result = BasicScanService().scan(
        [tmp_path], lambda done, total, text: progress.append((done, total, text))
    )

    assert result.source_count == 3
    assert result.image_count == 2
    assert result.spreadsheet_count == 1
    assert {finding.rule_id for finding in result.findings} == {
        "image.file.sha256",
        "excel.value.exact",
        "excel.row.exact",
        "excel.cell.target_operation",
    }
    assert result.algorithm_version == (
        "generic-image-local-1+western-blot-2+dot-blot-2+fluorescence-1+pathology-2+"
        "panel-split-1+excel-advanced-4"
    )
    assert result.completed_at is not None
    assert result.performance is not None
    assert result.performance.selected_backend == "cpu"
    assert {stage.stage_id for stage in result.performance.stages} >= {
        "source.collect",
        "image.generic_features",
        "spreadsheet.read",
        "result.sort",
    }
    assert progress[-1][:2] == (3, 3)


def test_corrupt_workbook_does_not_abort_other_files(tmp_path: Path) -> None:
    first = tmp_path / "one.png"
    second = tmp_path / "two.png"
    corrupt = tmp_path / "broken.xlsx"
    image = np.arange(32 * 32, dtype=np.uint8).reshape(32, 32)
    assert cv2.imwrite(str(first), image)
    second.write_bytes(first.read_bytes())
    corrupt.write_bytes(b"not-an-excel-archive")

    result = BasicScanService().scan([tmp_path])

    assert any(finding.rule_id == "image.file.sha256" for finding in result.findings)
    assert len(result.issues) == 1
    assert result.issues[0].source_path == str(corrupt)
    assert first.read_bytes() == second.read_bytes()
    assert corrupt.read_bytes() == b"not-an-excel-archive"


def test_basic_scan_mode_filters_files_and_skips_other_detector(tmp_path: Path) -> None:
    image = np.arange(32 * 32, dtype=np.uint8).reshape(32, 32)
    first_image = tmp_path / "one.png"
    assert cv2.imwrite(str(first_image), image)
    (tmp_path / "two.png").write_bytes(first_image.read_bytes())

    workbook = Workbook()
    workbook.active.append([2.5, 3.5])
    workbook.active.append([2.5, 3.5])
    workbook.save(tmp_path / "data.xlsx")

    image_result = BasicScanService(scan_mode=ScanMode.IMAGE).scan([tmp_path])
    data_result = BasicScanService(scan_mode=ScanMode.DATA).scan([tmp_path])

    assert (
        image_result.source_count,
        image_result.image_count,
        image_result.spreadsheet_count,
    ) == (
        2,
        2,
        0,
    )
    assert image_result.findings
    assert all(finding.rule_id.startswith("image.") for finding in image_result.findings)
    assert (data_result.source_count, data_result.image_count, data_result.spreadsheet_count) == (
        1,
        0,
        1,
    )
    assert data_result.findings
    assert all(finding.rule_id.startswith("excel.") for finding in data_result.findings)


def test_basic_scan_scans_only_selected_panels_and_returns_original_paths(tmp_path: Path) -> None:
    repeated = np.arange(80 * 100, dtype=np.uint8).reshape(80, 100)
    first = np.full((120, 240), 255, dtype=np.uint8)
    second = np.full((120, 240), 230, dtype=np.uint8)
    first[20:100, 10:110] = repeated
    second[20:100, 120:220] = repeated
    first_path = tmp_path / "first.png"
    second_path = tmp_path / "second.png"
    assert cv2.imwrite(str(first_path), first)
    assert cv2.imwrite(str(second_path), second)
    selections = (
        PanelSelection(str(first_path), 1, 1, 10, 20, 100, 80),
        PanelSelection(str(second_path), 1, 2, 120, 20, 100, 80),
    )

    result = BasicScanService(
        panel_splitting_enabled=True,
        panel_selections=selections,
        scan_mode=ScanMode.IMAGE,
    ).scan([first_path, second_path])

    exact = next(
        finding for finding in result.findings if finding.rule_id == "image.panel.pixel_exact"
    )
    assert exact.title == "子面板像素完全一致"
    assert {location.source_path for location in exact.locations} == {
        str(first_path.resolve()),
        str(second_path.resolve()),
    }
    assert all("子面板" in (location.coordinate or "") for location in exact.locations)
    assert exact.details["first_region_x"] in {10, 120}
    assert exact.details["first_region_width"] == 100
    assert result.performance is not None
    assert "image.panel_materialize" in {stage.stage_id for stage in result.performance.stages}


def test_panel_scan_reports_corrupt_image_without_aborting_valid_panels(tmp_path: Path) -> None:
    valid = tmp_path / "valid.png"
    corrupt = tmp_path / "corrupt.png"
    assert cv2.imwrite(str(valid), np.arange(32 * 32, dtype=np.uint8).reshape(32, 32))
    corrupt.write_bytes(b"not-an-image")
    selections = (
        PanelSelection(str(valid), 1, 1, 0, 0, 32, 32),
        PanelSelection(str(corrupt), 1, 1, 0, 0, 1, 1),
    )

    result = BasicScanService(
        panel_splitting_enabled=True,
        panel_selections=selections,
        scan_mode=ScanMode.IMAGE,
    ).scan([valid, corrupt])

    assert len(result.issues) == 1
    assert result.issues[0].source_path == str(corrupt.resolve())
    assert "复合图拆分无法读取图片" in result.issues[0].message


def test_basic_scan_uses_configured_digit_fragment_length(tmp_path: Path) -> None:
    path = tmp_path / "digit-fragments.xlsx"
    workbook = Workbook()
    workbook.active.append([14617])
    workbook.active.append([94617])
    workbook.save(path)

    default_result = BasicScanService().scan([path])
    stricter_result = BasicScanService(minimum_digit_run=5).scan([path])

    assert any(item.rule_id == "excel.digit_fragment" for item in default_result.findings)
    assert not any(item.rule_id == "excel.digit_fragment" for item in stricter_result.findings)


def test_scan_control_pauses_resumes_and_cancels_at_checkpoints() -> None:
    control = ScanControl()
    checkpoint_completed = Event()
    control.pause()

    thread = Thread(target=lambda: (control.checkpoint(), checkpoint_completed.set()))
    thread.start()
    assert not checkpoint_completed.wait(timeout=0.05)

    control.resume()
    assert checkpoint_completed.wait(timeout=1.0)
    thread.join(timeout=1.0)
    assert not thread.is_alive()

    control.cancel()
    with pytest.raises(ScanCancelled):
        control.checkpoint()


def test_basic_scan_cancel_discards_incomplete_result_at_next_file(tmp_path: Path) -> None:
    image = np.arange(32 * 32, dtype=np.uint8).reshape(32, 32)
    for index in range(3):
        assert cv2.imwrite(str(tmp_path / f"{index}.png"), image + index)
    control = ScanControl()

    def cancel_after_first(completed: int, total: int, message: str) -> None:
        del total, message
        if completed == 1:
            control.cancel()

    with pytest.raises(ScanCancelled):
        BasicScanService().scan([tmp_path], cancel_after_first, control)
